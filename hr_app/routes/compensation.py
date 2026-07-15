import json
import io
import os
import csv
from datetime import datetime, date, timedelta
from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, send_file
from flask_login import login_required, current_user
from sqlalchemy import extract, func
from ..extensions import db
from ..models.compensation import PayrollProfile, PayrollComponent, PayrollRun, PayrollSlip, SalaryRevision
from ..models.user import User
from ..models.attendance import Attendance
from ..models.holiday import OvertimeAccount, PayrollAuditLog
from ..models.pf import ProvidentFundConfig, PFContribution, PFLedger
from ..models.communication import Notification, NotificationRecipient
from ..models.tax import IncomeTaxSlab
from ..models.loan import LoanAdvanceRequest, LoanRepayment
from ..config import Config

comp_bp = Blueprint("compensation", __name__, url_prefix="/compensation")


@comp_bp.route("/")
@login_required
def index():
    if current_user.is_admin():
        profiles = PayrollProfile.query.all()
        runs = PayrollRun.query.order_by(PayrollRun.run_date.desc()).limit(12).all()
        slabs = IncomeTaxSlab.query.filter_by(is_active=True).order_by(IncomeTaxSlab.min_income).all()
        return render_template("compensation/index.html", profiles=profiles, runs=runs, slabs=slabs)
    profile = PayrollProfile.query.filter_by(user_id=current_user.id).first()
    slips = PayrollSlip.query.filter_by(user_id=current_user.id).order_by(PayrollSlip.created_at.desc()).all()
    return render_template("compensation/employee.html", profile=profile, slips=slips)


# ── Income Tax Slabs ──

@comp_bp.route("/tax-settings", methods=["GET", "POST"])
@login_required
def tax_settings():
    if not current_user.is_admin():
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            slab = IncomeTaxSlab(
                min_income=float(request.form["min_income"]),
                max_income=float(request.form["max_income"]),
                rate_pct=float(request.form["rate_pct"]),
                fixed_amount=float(request.form.get("fixed_amount", 0)),
            )
            db.session.add(slab)
            flash("Tax slab added.", "success")
        elif action == "delete":
            slab = IncomeTaxSlab.query.get_or_404(int(request.form["slab_id"]))
            db.session.delete(slab)
            flash("Tax slab deleted.", "success")
        elif action == "seed_defaults":
            for s in [
                (0, 600000, 0, 0),
                (600000, 1200000, 5, 0),
                (1200000, 2200000, 15, 30000),
                (2200000, 3200000, 25, 180000),
                (3200000, 4100000, 30, 430000),
                (4100000, 999999999, 35, 700000),
            ]:
                if not IncomeTaxSlab.query.filter_by(min_income=s[0]).first():
                    db.session.add(IncomeTaxSlab(min_income=s[0], max_income=s[1], rate_pct=s[2], fixed_amount=s[3]))
            flash("Default tax slabs seeded.", "success")
        db.session.commit()
        return redirect(url_for("compensation.tax_settings"))
    slabs = IncomeTaxSlab.query.order_by(IncomeTaxSlab.min_income).all()
    return render_template("compensation/tax_settings.html", slabs=slabs)


# ── Create Payroll Profile ──

@comp_bp.route("/create-profile", methods=["GET", "POST"])
@login_required
def create_profile():
    if not current_user.is_admin():
        flash("Access denied.", "danger")
        return redirect(url_for("compensation.index"))
    employees = User.query.filter_by(is_active=True).all()
    if request.method == "POST":
        uid = request.form.get("user_id", type=int)
        basic = request.form.get("basic_salary", type=float)
        effective = request.form.get("effective_from")
        if not uid or not basic or not effective:
            flash("All fields required.", "danger")
            return redirect(url_for("compensation.create_profile"))
        if PayrollProfile.query.filter_by(user_id=uid).first():
            flash("Profile already exists for this employee.", "warning")
            return redirect(url_for("compensation.edit_profile", uid=uid))
        try:
            eff_date = datetime.strptime(effective, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            flash("Invalid date format.", "danger")
            return redirect(url_for("compensation.create_profile"))
        profile = PayrollProfile(user_id=uid, basic_salary=basic, effective_from=eff_date)
        db.session.add(profile)
        db.session.flush()
        for cname in ["Basic Salary", "House Rent", "Medical Allowance", "Conveyance"]:
            comp = PayrollComponent(profile_id=profile.id, name=cname, type="allowance",
                                    value=basic * 0.25 if cname == "House Rent" else (basic * 0.1 if cname in ("Medical Allowance", "Conveyance") else basic),
                                    is_taxable=True if cname != "Medical Allowance" else False)
            db.session.add(comp)
        db.session.commit()
        flash(f"Profile created for {User.query.get(uid).full_name}.", "success")
        return redirect(url_for("compensation.edit_profile", uid=uid))
    return render_template("compensation/create_profile.html", employees=employees)


# ── Employee Salary Package Setup ──

@comp_bp.route("/edit-profile/<int:uid>", methods=["GET", "POST"])
@login_required
def edit_profile(uid):
    if not current_user.is_admin():
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))
    profile = PayrollProfile.query.filter_by(user_id=uid).first()
    emp = User.query.get_or_404(uid)
    if request.method == "POST":
        if not profile:
            profile = PayrollProfile(user_id=uid, basic_salary=0, effective_from=date.today())
            db.session.add(profile)
            db.session.flush()
        profile.basic_salary = float(request.form["basic_salary"])
        profile.effective_from = datetime.strptime(request.form["effective_from"], "%Y-%m-%d").date()
        submitted = json.loads(request.form.get("components_json", "[]"))
        for old in profile.components.all():
            db.session.delete(old)
        for comp in submitted:
            db.session.add(PayrollComponent(
                profile_id=profile.id, name=comp["name"],
                type=comp["type"],
                calculation_method=comp.get("method", "fixed"),
                value=float(comp["value"]),
                is_taxable=comp.get("is_taxable", True),
            ))
        db.session.commit()
        flash(f"Salary package updated for {emp.full_name}.", "success")
        return redirect(url_for("compensation.index"))
    return render_template("compensation/edit_profile.html", profile=profile, employee=emp)


# ── Payroll Run with Preview, Replace, Template Upload ──

def _preview_employee(pp, month, year, pf_config, adjustments=None):
    """Compute payroll preview for a single employee. Returns dict of computed values."""
    adj = adjustments.get(str(pp.user_id), {}) if adjustments else {}
    hire_date = pp.user.date_of_joining
    days_in_month = 30
    worked_days = days_in_month
    if hire_date and (hire_date.year > year or (hire_date.year == year and hire_date.month > month)):
        return None
    if hire_date and hire_date.month == month and hire_date.year == year:
        worked_days = days_in_month - hire_date.day + 1
    pro_ratio = worked_days / days_in_month
    basic = round(pp.basic_salary * pro_ratio, 2)
    allowances = 0; deductions = 0
    comps = {"allowances": {}, "deductions": {}}
    allowances_override = adj.get("allowances_override")
    for c in pp.components:
        val = round(c.value * pro_ratio, 2)
        if c.type == "allowance":
            if allowances_override is None:
                allowances += val
                comps["allowances"][c.name] = val
            else:
                comps["allowances"][c.name + " (override)"] = 0
        else:
            deductions += val
            comps["deductions"][c.name] = val
    if allowances_override is not None:
        allowances = float(allowances_override)
    ot_hours = db.session.query(func.sum(OvertimeAccount.overtime_hours)).filter(
        OvertimeAccount.user_id == pp.user_id,
        extract("month", OvertimeAccount.date) == month,
        extract("year", OvertimeAccount.date) == year
    ).scalar() or 0
    ot_pay = round(float(ot_hours) * (basic / max(160, 1)) * 1.5, 2)
    if allowances_override is None:
        allowances += ot_pay
    bonus_override = adj.get("bonus")
    if bonus_override:
        allowances += float(bonus_override)
    gross = round(basic + allowances, 2)
    monthly_tax = 0
    annual_gross = gross * 12
    tax_override = adj.get("income_tax")
    if tax_override is not None and tax_override != "":
        monthly_tax = round(float(tax_override), 2)
    else:
        annual_tax = IncomeTaxSlab.calculate_tax(annual_gross)
        monthly_tax = round(annual_tax / 12, 2)
    deductions += monthly_tax
    pf_employee = 0; pf_employer = 0
    if pf_config:
        pf_employee = round(gross * pf_config.employee_contribution_pct / 100, 2)
        pf_employer = round(gross * pf_config.employer_contribution_pct / 100, 2)
        deductions += pf_employee
    loan_deduction = 0
    active_loans = LoanAdvanceRequest.query.filter_by(user_id=pp.user_id, status="approved").filter(
        LoanAdvanceRequest.remaining_amount > 0).all()
    loan_details = []
    for loan in active_loans:
        if loan.remaining_amount and loan.remaining_amount > 0:
            installment = adj.get(f"loan_{loan.id}")
            ded = float(installment) if installment else (loan.monthly_installment or loan.amount / max(loan.installment_months, 1))
            ded = min(ded, loan.remaining_amount)
            if ded > 0:
                loan_deduction += ded
    deductions += loan_deduction
    custom_ded = adj.get("custom_deduction")
    if custom_ded:
        deductions += float(custom_ded)
    net = round(gross - deductions, 2)
    allowances_breakdown = dict(comps["allowances"])
    deductions_breakdown = dict(comps["deductions"])
    deductions_breakdown["Income Tax"] = round(monthly_tax, 2)
    if pf_config:
        deductions_breakdown["PF Employee"] = round(pf_employee, 2)
    if loan_deduction > 0:
        deductions_breakdown["Loan Repayment"] = round(loan_deduction, 2)
    if custom_ded:
        deductions_breakdown["Additional Deduction"] = round(float(custom_ded), 2)
    return {
        "user_id": pp.user_id,
        "name": pp.user.full_name,
        "designation": pp.user.designation or "",
        "basic": basic,
        "allowances_breakdown": allowances_breakdown,
        "allowances_total": round(allowances, 2),
        "bonus": float(adj.get("bonus", 0)) if adj.get("bonus") else 0,
        "ot_pay": ot_pay,
        "gross": gross,
        "monthly_tax": monthly_tax,
        "pf_employee": pf_employee,
        "loan_deduction": loan_deduction,
        "custom_ded": float(custom_ded) if custom_ded else 0,
        "deductions_total": round(deductions, 2),
        "net": net,
        "active_loans": active_loans,
    }


@comp_bp.route("/run-payroll", methods=["GET", "POST"])
@login_required
def run_payroll():
    if not current_user.is_admin():
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))
    pf_config = ProvidentFundConfig.query.first()

    if request.method == "POST":
        month = int(request.form["month"])
        year = int(request.form["year"])
        replace_all = request.form.get("replace_all") == "1"
        replace_ids_str = request.form.get("replace_ids", "[]")
        import json
        replace_ids = json.loads(replace_ids_str) if replace_ids_str else []

        existing = PayrollRun.query.filter_by(month=month, year=year).first()
        if existing:
            if replace_all:
                for slip in existing.slips:
                    for lr in LoanRepayment.query.filter_by(payroll_run_id=existing.id).all():
                        db.session.delete(lr)
                    PFContribution.query.filter_by(month=month, year=year).delete()
                    PFLedger.query.filter_by(transaction_date=date.today(),
                                              description=f"PF contribution {month}/{year}").delete()
                    db.session.delete(slip)
                db.session.delete(existing)
                db.session.commit()
                flash(f"Replaced existing payroll for {month}/{year}.", "info")
            else:
                flash(f"Payroll already run for {month}/{year}. Use 'Replace All' to override.", "danger")
                return redirect(url_for("compensation.run_payroll"))

        profiles = PayrollProfile.query.all()
        if not profiles:
            flash("No payroll profiles found.", "danger")
            return redirect(url_for("compensation.index"))
        adjustments = json.loads(request.form.get("adjustments", "{}"))

        pr = PayrollRun(month=month, year=year, processed_by=current_user.id, status="processing")
        db.session.add(pr)
        db.session.flush()
        total_gross = 0; total_ded = 0; total_net = 0; emp_count = 0
        total_tax = 0; total_pf_ee = 0; total_pf_er = 0; total_loan = 0; total_custom = 0

        for pp in profiles:
            if not pp.user.is_active:
                continue
            uid_str = str(pp.user_id)
            if replace_ids and uid_str not in replace_ids and not replace_all:
                continue
            preview = _preview_employee(pp, month, year, pf_config, adjustments)
            if preview is None:
                continue
            basic = preview["basic"]
            allowances = preview["allowances_total"]
            deductions = preview["deductions_total"]
            gross = preview["gross"]
            net = preview["net"]
            monthly_tax = preview["monthly_tax"]
            pf_employee = preview["pf_employee"]
            loan_deduction = preview["loan_deduction"]
            custom_ded = preview["custom_ded"]
            pf_employer = round(gross * pf_config.employer_contribution_pct / 100, 2) if pf_config else 0

            comps = {"allowances": preview["allowances_breakdown"], "deductions": {}}
            for c in pp.components:
                if c.type != "allowance":
                    val = round(c.value * (basic / pp.basic_salary if pp.basic_salary else 1), 2)
                    comps["deductions"][c.name] = val
            comps["deductions"]["Income Tax"] = round(monthly_tax, 2)
            if pf_config:
                comps["deductions"]["PF Employee"] = round(pf_employee, 2)
            if loan_deduction > 0:
                comps["deductions"]["Loan Repayment"] = round(loan_deduction, 2)
            if custom_ded > 0:
                comps["deductions"]["Additional Deduction"] = round(custom_ded, 2)

            adj = adjustments.get(uid_str, {})
            active_loans = LoanAdvanceRequest.query.filter_by(user_id=pp.user_id, status="approved").filter(
                LoanAdvanceRequest.remaining_amount > 0).all()
            for loan in active_loans:
                if loan.remaining_amount and loan.remaining_amount > 0:
                    installment = adj.get(f"loan_{loan.id}")
                    ded = float(installment) if installment else (loan.monthly_installment or loan.amount / max(loan.installment_months, 1))
                    ded = min(ded, loan.remaining_amount)
                    if ded > 0:
                        loan.remaining_amount -= ded
                        if loan.remaining_amount <= 0:
                            loan.status = "paid"
                        db.session.add(LoanRepayment(loan_id=loan.id, amount=ded, payroll_run_id=pr.id,
                                                      notes=f"Deducted via payroll {month}/{year}"))

            slip = PayrollSlip(
                payroll_run_id=pr.id, user_id=pp.user_id, basic_salary=basic,
                allowances=allowances, deductions=deductions,
                gross_pay=gross, total_deductions=deductions,
                net_pay=net, components_json=json.dumps(comps),
            )
            db.session.add(slip)
            total_gross += gross; total_ded += deductions; total_net += net; emp_count += 1

            total_tax += monthly_tax
            total_pf_ee += pf_employee
            total_pf_er += pf_employer
            total_loan += loan_deduction
            total_custom += custom_ded

            if pf_config:
                contrib = PFContribution(user_id=pp.user_id, month=month, year=year,
                                          employee_amount=pf_employee, employer_amount=pf_employer,
                                          total_amount=round(pf_employee + pf_employer, 2))
                db.session.add(contrib)
                db.session.add(PFLedger(user_id=pp.user_id, transaction_date=date.today(),
                                         transaction_type="contribution",
                                         description=f"PF contribution {month}/{year}",
                                         credit=round(pf_employee + pf_employer, 2), debit=0))
            _notify(pp.user_id, "Salary Slip Generated",
                    f"Your salary slip for {month}/{year} is available.", "info", "compensation")
            if pp.user.manager_id:
                _notify(pp.user.manager_id, f"Salary Slip: {pp.user.full_name}",
                        f"Salary for {month}/{year}: Rs.{net:,.0f}", "info", "compensation")

        pr.status = "completed"
        pr.total_gross = round(total_gross, 2)
        pr.total_deductions = round(total_ded, 2)
        pr.total_net = round(total_net, 2)
        pr.employee_count = emp_count

        # Post salary journal entry to general ledger
        from shared.ledger_utils import get_or_create_account, post_journal_entry
        salary_exp = get_or_create_account("5121", "Salary Expense", "expense", parent_code="512")
        salary_payable = get_or_create_account("2121", "Salary Payable", "liability", parent_code="212")
        tax_payable = get_or_create_account("2122", "Income Tax Payable", "liability", parent_code="212")
        pf_payable = get_or_create_account("2123", "PF Payable", "liability", parent_code="212")
        loan_clearing = get_or_create_account("2124", "Loan Deductions Clearing", "liability", parent_code="212")
        pf_exp = get_or_create_account("5122", "PF Employer Expense", "expense", parent_code="512")

        lines = [{"account_id": salary_exp.id, "debit": round(total_gross, 2), "credit": 0,
                   "description": f"Gross salary {month}/{year}"}]
        if round(total_tax, 2) > 0:
            lines.append({"account_id": tax_payable.id, "debit": 0, "credit": round(total_tax, 2),
                           "description": f"Income tax deducted"})
        if round(total_pf_ee, 2) > 0:
            lines.append({"account_id": pf_payable.id, "debit": 0, "credit": round(total_pf_ee, 2),
                           "description": f"PF employee contribution"})
        if round(total_loan, 2) > 0:
            lines.append({"account_id": loan_clearing.id, "debit": 0, "credit": round(total_loan, 2),
                           "description": f"Loan repayment deductions"})
        if round(total_custom, 2) > 0:
            lines.append({"account_id": salary_payable.id, "debit": 0, "credit": round(total_custom, 2),
                           "description": f"Custom deductions"})
        lines.append({"account_id": salary_payable.id, "debit": 0, "credit": round(total_net, 2),
                       "description": f"Net salary payable"})
        post_journal_entry(voucher_type="PRL", voucher_id=pr.id,
                           voucher_number=f"PRL-{year}{month:02d}",
                           description=f"Payroll Run {month}/{year}",
                           lines=lines, entry_date=datetime.utcnow(),
                           created_by=current_user.id)

        if round(total_pf_er, 2) > 0:
            pf_lines = [{"account_id": pf_exp.id, "debit": round(total_pf_er, 2), "credit": 0,
                          "description": f"PF employer contribution {month}/{year}"},
                        {"account_id": pf_payable.id, "debit": 0, "credit": round(total_pf_er, 2),
                          "description": f"PF employer contribution"}]
            post_journal_entry(voucher_type="PRL", voucher_id=pr.id,
                               voucher_number=f"PRL-{year}{month:02d}-PF",
                               description=f"PF Employer Contribution {month}/{year}",
                               lines=pf_lines, entry_date=datetime.utcnow(),
                               created_by=current_user.id)

        db.session.add(PayrollAuditLog(payroll_run_id=pr.id, action="payroll_run",
                                        performed_by=current_user.id, details=f"Run for {month}/{year}",
                                        ip_address=request.remote_addr))
        db.session.commit()
        flash(f"Payroll completed for {emp_count} employees. Net: Rs.{total_net:,.0f}", "success")
        return redirect(url_for("compensation.index"))

    # GET: build preview for all employees
    preview_employees = []
    profiles = PayrollProfile.query.all()
    for pp in profiles:
        if not pp.user.is_active:
            continue
        p = _preview_employee(pp, 7, 2026, pf_config)  # default preview month
        if p:
            active_loans = LoanAdvanceRequest.query.filter_by(user_id=pp.user_id, status="approved").filter(
                LoanAdvanceRequest.remaining_amount > 0).all()
            p["active_loans"] = active_loans
            preview_employees.append(p)
    return render_template("compensation/run_payroll.html", employees=preview_employees,
                           pf_config=pf_config, now=datetime.utcnow())


@comp_bp.route("/payroll-preview-json", methods=["POST"])
@login_required
def payroll_preview_json():
    """Return computed preview as JSON for AJAX recalculation."""
    if not current_user.is_admin():
        return jsonify({"error": "Access denied"}), 403
    month = int(request.form.get("month", 7))
    year = int(request.form.get("year", 2026))
    adjustments = json.loads(request.form.get("adjustments", "{}"))
    pf_config = ProvidentFundConfig.query.first()
    results = []
    for pp in PayrollProfile.query.all():
        if not pp.user.is_active:
            continue
        p = _preview_employee(pp, month, year, pf_config, adjustments)
        if p:
            results.append(p)
    return jsonify({"employees": results})


@comp_bp.route("/download-template")
@login_required
def download_payroll_template():
    """Download an Excel template for bulk payroll data upload."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash("openpyxl not installed.", "danger")
        return redirect(url_for("compensation.index"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll Data"
    headers = [
        "Employee Code", "Full Name", "Basic Salary", "Allowances", "Bonus",
        "Income Tax", "Loan Deduction", "Custom Deduction", "Net Pay"
    ]
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = thin
    sample = ["EMP001", "John Doe", 50000, 25000, 5000, 2500, 2000, 0, 73500]
    for ci, val in enumerate(sample, 1):
        c = ws.cell(row=2, column=ci, value=val)
        c.border = thin; c.alignment = ha
    from openpyxl.utils import get_column_letter
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 20
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="payroll_template.xlsx")


@comp_bp.route("/upload-bulk-data", methods=["POST"])
@login_required
def upload_bulk_data():
    """Upload bulk payroll data as XLSX/CSV/JSON and apply as adjustments."""
    if not current_user.is_admin():
        return jsonify({"error": "Access denied"}), 403
    adjustments = {}

    if request.content_type and "multipart/form-data" in request.content_type:
        f = request.files.get("file")
        if not f or f.filename == "":
            return jsonify({"error": "No file uploaded"}), 400
        ext = os.path.splitext(f.filename)[1].lower()
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(Config.UPLOAD_FOLDER, f"payroll_bulk_{ts}{ext}")
        f.save(path)
        rows = []
        try:
            if ext == ".csv":
                import csv
                with open(path, newline="", encoding="utf-8-sig") as fh:
                    reader = csv.DictReader(fh)
                    rows = [r for r in reader]
            elif ext in (".xlsx", ".xls"):
                import openpyxl
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if all(v is None or str(v).strip() == "" for v in row):
                        continue
                    rows.append(dict(zip(headers, [str(v).strip() if v is not None else "" for v in row])))
                wb.close()
        except Exception as e:
            return jsonify({"error": f"Parse error: {str(e)}"}), 400
        for row in rows:
            code = row.get("Employee Code", row.get("employee_code", "")).strip()
            user = User.query.filter_by(employee_code=code).first()
            if not user:
                user = User.query.filter_by(email=code).first()
            if not user:
                continue
            uid = str(user.id)
            adj = {}
            bonus = row.get("Bonus") or row.get("bonus")
            if bonus: adj["bonus"] = bonus
            tax = row.get("Income Tax") or row.get("income_tax") or row.get("tax")
            if tax: adj["income_tax"] = tax
            custom = row.get("Custom Deduction") or row.get("custom_deduction") or row.get("extra_ded")
            if custom: adj["custom_deduction"] = custom
            if adj:
                adjustments[uid] = adj
        msg = f"Bulk data loaded for {len(adjustments)} employees from {f.filename}."
        flash(msg, "success")
        return jsonify({"adjustments": adjustments, "message": msg})

    # JSON upload
    data = request.get_json(silent=True)
    if not data:
        flash("Invalid upload data.", "danger")
        return redirect(url_for("compensation.run_payroll"))
    payroll_data = data.get("data", [])
    for row in payroll_data:
        code = row.get("employee_code", "").strip()
        user = User.query.filter_by(employee_code=code).first()
        if not user:
            continue
        uid = str(user.id)
        adjustments[uid] = {
            "bonus": row.get("bonus"),
            "income_tax": row.get("income_tax"),
            "custom_deduction": row.get("custom_deduction"),
        }
    flash(f"Bulk data loaded for {len(adjustments)} employees. Adjustments applied.", "success")
    return jsonify({"adjustments": adjustments})


# ── View & Download Slip ──

@comp_bp.route("/slip/<int:sid>")
@login_required
def view_slip(sid):
    slip = PayrollSlip.query.get_or_404(sid)
    if slip.user_id != current_user.id and not current_user.is_admin():
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))
    comps = _normalize_comps(json.loads(slip.components_json)) if slip.components_json else {"allowances": {}, "deductions": {}}
    return render_template("compensation/slip.html", slip=slip, comps=comps)


@comp_bp.route("/slip/<int:sid>/pdf")
@login_required
def download_slip_pdf(sid):
    slip = PayrollSlip.query.get_or_404(sid)
    if slip.user_id != current_user.id and not current_user.is_admin():
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
    except ImportError:
        flash("PDF library not installed.", "danger")
        return redirect(url_for("compensation.index"))
    comps = _normalize_comps(json.loads(slip.components_json)) if slip.components_json else {"allowances": {}, "deductions": {}}
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    pw, ph = A4
    margin = 50
    col_mid = pw / 2

    # ── Header ──
    c.setFillColor(colors.HexColor("#1a237e"))
    c.setFont("Helvetica-Bold", 20)
    c.drawCentredString(pw / 2, ph - 40, "Solarkon (Private) Limited")
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 11)
    c.drawCentredString(pw / 2, ph - 58, f"Salary Slip - {slip.payroll_run.month}/{slip.payroll_run.year}")
    c.setStrokeColor(colors.HexColor("#1a237e"))
    c.setLineWidth(2)
    c.line(margin, ph - 65, pw - margin, ph - 65)

    # ── Employee Info ──
    y = ph - 88
    c.setFont("Helvetica", 9.5)
    c.drawString(margin, y, f"Employee: {slip.user.full_name}")
    c.drawString(col_mid + 10, y, f"Designation: {slip.user.designation or '-'}")
    y -= 15
    c.drawString(margin, y, f"Department: {slip.user.department or '-'}")
    c.drawString(col_mid + 10, y, f"Code: {slip.user.employee_code}")
    y -= 15
    c.drawString(margin, y, f"CNIC: {slip.user.cnic or '-'}")
    y -= 22

    # ── Earnings & Deductions Header ──
    c.setFillColor(colors.HexColor("#1a237e"))
    c.setFont("Helvetica-Bold", 11)
    c.drawString(margin, y, "Earnings")
    c.drawString(col_mid + 10, y, "Deductions")
    c.setFillColor(colors.black)
    y -= 16

    # Build lists
    earnings = [("Basic Salary", slip.basic_salary)]
    for name, val in comps.get("allowances", {}).items():
        if val:
            earnings.append((name, val))
    deductions_list = []
    for name, val in comps.get("deductions", {}).items():
        if val:
            deductions_list.append((name, val))

    # Draw earnings/deductions rows with alternating shading
    max_rows = max(len(earnings), len(deductions_list), 0)
    c.setFont("Helvetica", 9.5)
    for i in range(max_rows):
        if i % 2 == 1:
            c.setFillColor(colors.HexColor("#f0f4ff"))
        else:
            c.setFillColor(colors.white)
        c.rect(margin, y - 2, col_mid - margin - 5, 14, fill=1, stroke=0)
        c.rect(col_mid + 5, y - 2, pw - margin - col_mid - 5, 14, fill=1, stroke=0)
        c.setFillColor(colors.black)
        if i < len(earnings):
            label, val = earnings[i]
            c.drawString(margin + 3, y, str(label))
            c.drawRightString(col_mid - 10, y, f"Rs. {val:,.2f}")
        if i < len(deductions_list):
            label, val = deductions_list[i]
            c.drawString(col_mid + 10, y, str(label))
            c.drawRightString(pw - margin, y, f"Rs. {val:,.2f}")
        y -= 14
    y -= 6

    # ── Totals ──
    c.setStrokeColor(colors.HexColor("#1a237e"))
    c.setLineWidth(1)
    c.line(margin, y, pw - margin, y)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(margin, y - 14, "Total Earnings")
    c.drawRightString(col_mid - 10, y - 14, f"Rs. {slip.gross_pay:,.2f}")
    c.drawString(col_mid + 10, y - 14, "Total Deductions")
    c.drawRightString(pw - margin, y - 14, f"Rs. {slip.total_deductions:,.2f}")
    y -= 24

    # ── Net Pay Box ──
    box_y = y - 50
    c.setStrokeColor(colors.HexColor("#1b5e20"))
    c.setFillColor(colors.HexColor("#e8f5e9"))
    c.setLineWidth(2)
    c.roundRect(margin, box_y, pw - 2 * margin, 48, 6, fill=1, stroke=1)
    c.setFont("Helvetica-Bold", 14)
    c.setFillColor(colors.HexColor("#1b5e20"))
    c.drawString(margin + 12, box_y + 16, f"Net Payable: Rs. {int(slip.net_pay):,}")
    # Amount in words (wrapped if too long)
    words = _num_to_words(int(slip.net_pay))
    c.setFont("Helvetica", 9)
    c.drawRightString(pw - margin - 12, box_y + 18, f"(Rupees {words})")
    y = box_y - 20

    # ── Footer ──
    c.setFont("Helvetica", 7.5)
    c.setFillColor(colors.HexColor("#9ca3af"))
    c.drawCentredString(pw / 2, 30, f"Generated on {datetime.now():%d-%b-%Y %H:%M} | This is a computer-generated slip")

    c.save()
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf", as_attachment=False,
                     download_name=f"salary_slip_{slip.payroll_run.month}_{slip.payroll_run.year}.pdf")


def _normalize_comps(comps):
    """Normalize components_json to {allowances: {}, deductions: {}} regardless of storage format."""
    if isinstance(comps, dict):
        if "allowances" in comps and "deductions" in comps:
            return comps
        result = {"allowances": {}, "deductions": {}}
        for k, v in comps.items():
            result["allowances"][k] = float(v) if v else 0
        return result
    if isinstance(comps, list):
        result = {"allowances": {}, "deductions": {}}
        for item in comps:
            if isinstance(item, dict):
                name = item.get("name", "Item")
                val = float(item.get("value", 0))
                typ = item.get("type", "allowance")
                if typ in ("deduction", "tax", "loan"):
                    result["deductions"][name] = val
                else:
                    result["allowances"][name] = val
        return result
    return {"allowances": {}, "deductions": {}}


def _num_to_words(n):
    if n == 0: return "Zero"
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
            "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
            "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    def _convert(num):
        if num < 20: return ones[num]
        if num < 100: return tens[num // 10] + (" " + ones[num % 10] if num % 10 else "")
        if num < 1000: return ones[num // 100] + " Hundred" + (" " + _convert(num % 100) if num % 100 else "")
        if num < 100000: return _convert(num // 1000) + " Thousand" + (" " + _convert(num % 1000) if num % 1000 else "")
        if num < 10000000: return _convert(num // 100000) + " Lakh" + (" " + _convert(num % 100000) if num % 100000 else "")
        return _convert(num // 10000000) + " Crore" + (" " + _convert(num % 10000000) if num % 10000000 else "")
    return _convert(n) + " Only"


def _notify(user_id, title, message, ntype="info", module="compensation", ref_id=None):
    notif = Notification(title=title, message=message, notification_type=ntype,
                         module=module, reference_id=ref_id, created_by=current_user.id)
    db.session.add(notif)
    db.session.flush()
    db.session.add(NotificationRecipient(notification_id=notif.id, user_id=user_id))


# ── Bank Ledger & Revisions ──

@comp_bp.route("/export-bank-ledger/<int:rid>")
@login_required
def export_bank_ledger(rid):
    if not current_user.is_admin():
        return jsonify({"error": "Access denied"}), 403
    run = PayrollRun.query.get_or_404(rid)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Payroll_{run.month}_{run.year}"
    headers = ["Employee Code", "Employee Name", "Bank Name", "Account Title", "Account Number", "Net Pay"]
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center")
    for ri, slip in enumerate(run.slips, 2):
        u = slip.user
        for ci, val in enumerate([u.employee_code, u.full_name, u.bank_name or "",
                                   u.bank_account_title or "", u.bank_account_number or "",
                                   slip.net_pay], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center")
    from openpyxl.utils import get_column_letter
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 22
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"bank_ledger_{run.month}_{run.year}.xlsx")


@comp_bp.route("/revisions")
@login_required
def revisions():
    if not current_user.is_admin():
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))
    all_revisions = SalaryRevision.query.order_by(SalaryRevision.created_at.desc()).all()
    return render_template("compensation/revisions.html", revisions=all_revisions)
