import io
import csv
from datetime import datetime, date, timedelta
from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required, current_user
from sqlalchemy import func, extract, case
from ..extensions import db
from ..models.user import User
from ..models.attendance import Attendance
from ..models.leave import LeaveRequest, LeaveType, LeaveQuota
from ..models.timesheet import TimesheetWeek, TimesheetEntry
from ..models.compensation import PayrollRun, PayrollSlip
from ..models.pf import PFLedger, PFContribution
from ..models.holiday import OvertimeAccount
from ..models.loan import LoanAdvanceRequest, LoanRepayment

reports_bp = Blueprint("reports", __name__, url_prefix="/reports")


def _require_admin():
    if not current_user.is_admin() and not current_user.is_manager():
        return False
    return True


def _get_scope_users(scope, user_id=None, department=None):
    if scope == "individual" and user_id:
        return [User.query.get(user_id)]
    elif scope == "department" and department:
        return User.query.filter_by(department=department, is_active=True).all()
    else:
        return User.query.filter_by(is_active=True).all()


@reports_bp.route("/")
@login_required
def index():
    if not _require_admin():
        return render_template("dashboard/index.html")
    employees = User.query.filter_by(is_active=True).all()
    return render_template("reports/index.html", employees=employees)


@reports_bp.route("/dashboard")
@login_required
def dashboard():
    today = date.today()
    if current_user.is_admin() or current_user.is_manager():
        total_employees = User.query.filter_by(is_active=True).count()
        today_present = Attendance.query.filter(Attendance.date == today, Attendance.clock_in != None).count()
        pending_leaves = LeaveRequest.query.filter(LeaveRequest.status == "pending").count()
        running_payroll = PayrollRun.query.filter(
            PayrollRun.status == "draft", extract("month", PayrollRun.run_date) == today.month
        ).first()
        return jsonify({
            "total_employees": total_employees,
            "today_present": today_present,
            "today_absent": total_employees - today_present,
            "pending_leaves": pending_leaves,
            "active_payroll": running_payroll is not None,
        })
    my_att = Attendance.query.filter(Attendance.user_id == current_user.id, Attendance.date == today).first()
    my_pending = LeaveRequest.query.filter(
        LeaveRequest.user_id == current_user.id, LeaveRequest.status == "pending"
    ).count()
    return jsonify({
        "total_employees": "-",
        "today_present": "Clocked In" if my_att and my_att.clock_in and not my_att.clock_out else "Done" if my_att and my_att.clock_out else "Not Clocked",
        "today_absent": "-",
        "pending_leaves": my_pending,
        "active_payroll": False,
    })


@reports_bp.route("/attendance-chart")
@login_required
def attendance_chart():
    if not _require_admin():
        return jsonify({"error": "Access denied"}), 403
    year = request.args.get("year", datetime.now().year, type=int)
    monthly = db.session.query(
        extract("month", Attendance.date).label("month"),
        func.count(Attendance.id).label("total"),
        func.sum(case((Attendance.is_late == True, 1), else_=0)).label("late"),
        func.sum(case((Attendance.is_half_day == True, 1), else_=0)).label("half"),
    ).filter(extract("year", Attendance.date) == year).group_by("month").order_by("month").all()
    return jsonify([{
        "month": int(r.month), "total": int(r.total),
        "late": int(r.late), "half": int(r.half),
    } for r in monthly])


@reports_bp.route("/query", methods=["POST"])
@login_required
def query():
    if not _require_admin():
        return jsonify({"error": "Access denied"}), 403
    data = request.get_json()
    scope = data.get("scope", "company")
    user_id = data.get("user_id")
    if user_id is not None:
        user_id = int(user_id)
    department = data.get("department")
    date_from = datetime.strptime(data["date_from"], "%Y-%m-%d").date() if data.get("date_from") else None
    date_to = datetime.strptime(data["date_to"], "%Y-%m-%d").date() if data.get("date_to") else None
    columns = data.get("columns", ["employee", "total_hours", "leaves", "sick_days"])
    users = _get_scope_users(scope, user_id, department)
    if scope == "individual" and current_user.is_manager() and user_id:
        emp = User.query.get(user_id)
        if emp and emp.manager_id != current_user.id and not current_user.is_admin():
            return jsonify({"error": "Not authorized for this employee"}), 403
    results = []
    for u in users:
        row = {"employee": u.full_name, "department": u.department, "id": u.id}
        base = Attendance.query.filter(Attendance.user_id == u.id)
        if date_from:
            base = base.filter(Attendance.date >= date_from)
        if date_to:
            base = base.filter(Attendance.date <= date_to)
        if "total_hours" in columns:
            totals = base.with_entities(
                func.sum(case((Attendance.clock_out != None, func.julianday(Attendance.clock_out) - func.julianday(Attendance.clock_in)), else_=0))
            ).scalar() or 0
            row["total_hours"] = round(float(totals) * 24, 2)
        if "overtime" in columns:
            ot = db.session.query(func.sum(OvertimeAccount.overtime_hours)).filter(
                OvertimeAccount.user_id == u.id
            ).scalar() or 0
            row["overtime"] = round(float(ot), 2)
        if "leaves" in columns:
            lq = LeaveRequest.query.filter(LeaveRequest.user_id == u.id, LeaveRequest.status == "approved")
            if date_from:
                lq = lq.filter(LeaveRequest.start_date >= date_from)
            if date_to:
                lq = lq.filter(LeaveRequest.end_date <= date_to)
            row["leaves"] = sum(l.total_days for l in lq.all())
        if "sick_days" in columns:
            sq = LeaveRequest.query.join(LeaveType).filter(
                LeaveRequest.user_id == u.id, LeaveRequest.status == "approved",
                LeaveType.code == "SL"
            )
            if date_from:
                sq = sq.filter(LeaveRequest.start_date >= date_from)
            if date_to:
                sq = sq.filter(LeaveRequest.end_date <= date_to)
            row["sick_days"] = sum(l.total_days for l in sq.all())
        if "late_days" in columns:
            late = base.filter(Attendance.is_late == True).count()
            row["late_days"] = late
        if "present_days" in columns:
            present = base.filter(Attendance.clock_in != None).count()
            row["present_days"] = present
        if "pf_balance" in columns:
            pf = db.session.query(func.sum(PFLedger.credit) - func.sum(PFLedger.debit)).filter(
                PFLedger.user_id == u.id
            ).scalar() or 0
            row["pf_balance"] = round(float(pf), 2)
        if "loan_balance" in columns:
            loan = db.session.query(func.sum(LoanAdvanceRequest.remaining_amount)).filter(
                LoanAdvanceRequest.user_id == u.id, LoanAdvanceRequest.status == "approved"
            ).scalar() or 0
            row["loan_balance"] = round(float(loan), 2)
        results.append(row)
    return jsonify({"results": results})


@reports_bp.route("/export-excel", methods=["POST"])
@login_required
def export_excel():
    if not _require_admin():
        return jsonify({"error": "Access denied"}), 403
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500
    data = request.get_json()
    rows = data.get("rows", [])
    columns = data.get("columns", [])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")
    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"hr_report_{date.today().isoformat()}.xlsx")


@reports_bp.route("/export-attendance")
@login_required
def export_attendance():
    if not _require_admin():
        return jsonify({"error": "Access denied"}), 403
    month = request.args.get("month", datetime.now().month, type=int)
    year = request.args.get("year", datetime.now().year, type=int)
    records = db.session.query(Attendance, User).join(User).filter(
        extract("month", Attendance.date) == month,
        extract("year", Attendance.date) == year,
    ).order_by(Attendance.date).all()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Date", "Employee", "Department", "Clock In", "Clock Out", "Status", "Late", "Half Day"])
    for att, usr in records:
        w.writerow([
            att.date, usr.full_name, usr.department,
            att.clock_in.strftime("%H:%M") if att.clock_in else "",
            att.clock_out.strftime("%H:%M") if att.clock_out else "",
            att.status, "Yes" if att.is_late else "No", "Yes" if att.is_half_day else "No"
        ])
    buf.seek(0)
    return send_file(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        mimetype="text/csv", as_attachment=True,
        download_name=f"attendance_{year}_{month:02d}.csv"
    )


@reports_bp.route("/export-leaves")
@login_required
def export_leaves():
    if not _require_admin():
        return jsonify({"error": "Access denied"}), 403
    year = request.args.get("year", datetime.now().year, type=int)
    records = db.session.query(LeaveRequest, User, LeaveType).join(User).join(LeaveType).filter(
        extract("year", LeaveRequest.start_date) == year
    ).order_by(LeaveRequest.start_date).all()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Employee", "Type", "Start", "End", "Days", "Status", "Reason"])
    for lr, usr, lt in records:
        w.writerow([usr.full_name, lt.name, lr.start_date, lr.end_date, lr.total_days, lr.status, lr.reason])
    buf.seek(0)
    return send_file(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        mimetype="text/csv", as_attachment=True,
        download_name=f"leaves_{year}.csv"
    )
