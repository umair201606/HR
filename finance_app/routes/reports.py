from datetime import datetime, date, timedelta
from decimal import Decimal
from flask import Blueprint, render_template, request, jsonify, Response, send_file
from flask_login import login_required
from sqlalchemy import text
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from collections import defaultdict
from shared.extensions import db
from shared.models.ledger import ChartOfAccount, JournalEntry, JournalLine
from shared.models.base import User
from shared.models.company_settings import AccountingPeriod, ReportSettings
from shared.ledger_utils import posting_account

finance_bp = Blueprint("finance", __name__, url_prefix="/finance")

ACCOUNT_TYPES = {"asset": "Asset", "liability": "Liability",
                 "equity": "Equity", "revenue": "Revenue", "expense": "Expense"}

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
SUBTITLE_FONT = Font(bold=True, size=10, color="555555")
DATA_FONT = Font(size=10)
BOLD_FONT = Font(bold=True, size=10)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


def _parse_date(d):
    if not d:
        return None
    if isinstance(d, date):
        return d
    for fmt in ("%Y-%m-%d", "%d/%m/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(d, fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _default_period():
    """The period to default reports to: the one containing today (preferring an
    active one), else the most recent active period. Robust even when several
    periods are erroneously marked active at once."""
    today = date.today()
    p = (AccountingPeriod.query
         .filter(AccountingPeriod.start_date <= today,
                 AccountingPeriod.end_date >= today)
         .order_by(AccountingPeriod.is_active.desc(),
                   AccountingPeriod.start_date.desc())
         .first())
    if p:
        return p
    return (AccountingPeriod.query.filter_by(is_active=True)
            .order_by(AccountingPeriod.start_date.desc()).first())


def _resolve_period():
    filter_mode = request.args.get("filter_mode", "period")
    period_id = request.args.get("period_id", type=int)
    from_str = request.args.get("from", "").strip()
    to_str = request.args.get("to", "").strip()
    from_date = _parse_date(from_str) if from_str else None
    to_date = _parse_date(to_str) if to_str else None

    periods = AccountingPeriod.query.order_by(AccountingPeriod.start_date.desc()).all()
    selected_period_id = period_id

    if filter_mode == "period" and period_id:
        period = AccountingPeriod.query.get(period_id)
        if period:
            from_date = period.start_date
            to_date = period.end_date

    if not from_date and not to_date:
        active = _default_period()
        if active:
            from_date = active.start_date
            to_date = active.end_date
            if not selected_period_id:
                selected_period_id = active.id
    elif not selected_period_id and filter_mode == "period":
        active = _default_period()
        if active:
            selected_period_id = active.id

    return from_date, to_date, periods, selected_period_id, filter_mode, from_str, to_str


def _get_account_balance(account_id, as_of=None):
    q = db.session.query(
        db.func.coalesce(db.func.sum(JournalLine.debit), 0).label("dr"),
        db.func.coalesce(db.func.sum(JournalLine.credit), 0).label("cr"),
    ).join(JournalEntry, JournalLine.journal_entry_id == JournalEntry.id
           ).filter(JournalLine.account_id == account_id,
                    JournalEntry.is_posted == True)
    if as_of:
        q = q.filter(JournalEntry.entry_date <= as_of)
    row = q.first()
    return Decimal(str(row.dr)), Decimal(str(row.cr))


def _all_account_balances(as_of=None, account_types=None):
    q = db.session.query(
        JournalLine.account_id,
        ChartOfAccount.code,
        ChartOfAccount.name,
        ChartOfAccount.type,
        db.func.coalesce(db.func.sum(JournalLine.debit), 0).label("dr"),
        db.func.coalesce(db.func.sum(JournalLine.credit), 0).label("cr"),
    ).join(JournalEntry, JournalLine.journal_entry_id == JournalEntry.id
           ).join(ChartOfAccount, JournalLine.account_id == ChartOfAccount.id
                  ).filter(JournalEntry.is_posted == True)
    if as_of:
        q = q.filter(JournalEntry.entry_date <= as_of)
    if account_types:
        q = q.filter(ChartOfAccount.type.in_(account_types))
    q = q.group_by(JournalLine.account_id, ChartOfAccount.code,
                   ChartOfAccount.name, ChartOfAccount.type
                   ).order_by(ChartOfAccount.code)
    return q.all()


def _net_income(as_of=None):
    rev = _all_account_balances(as_of, ["revenue"])
    exp = _all_account_balances(as_of, ["expense"])
    total_rev = sum((r.cr - r.dr) for r in rev) if rev else Decimal("0")
    total_exp = sum((e.dr - e.cr) for e in exp) if exp else Decimal("0")
    return total_rev - total_exp


def _period_movements(from_date=None, to_date=None, types=None):
    """Per-account (dr, cr) sums of posted lines within the period.
    Returns {account_id: (Decimal dr, Decimal cr)}."""
    q = db.session.query(
        JournalLine.account_id,
        db.func.coalesce(db.func.sum(JournalLine.debit), 0).label("dr"),
        db.func.coalesce(db.func.sum(JournalLine.credit), 0).label("cr"),
    ).join(JournalEntry, JournalLine.journal_entry_id == JournalEntry.id
           ).filter(JournalEntry.is_posted == True)
    if from_date:
        q = q.filter(JournalEntry.entry_date >= from_date)
    if to_date:
        q = q.filter(JournalEntry.entry_date <= to_date)
    if types:
        q = q.join(ChartOfAccount, JournalLine.account_id == ChartOfAccount.id
                   ).filter(ChartOfAccount.type.in_(types))
    return {r.account_id: (Decimal(str(r.dr)), Decimal(str(r.cr)))
            for r in q.group_by(JournalLine.account_id).all()}


def _pl_rows(from_date, to_date):
    """Sectioned P&L per ReportSettings.pl_structure.

    Every P&L account's contribution to profit is (credit - debit); revenue
    is naturally positive, expenses negative, and contra accounts (sales
    returns, purchase discounts) self-correct without special cases. Each
    structure entry's ``negate`` flag only flips the DISPLAY sign so expense
    sections read as positive figures under a "Less: ..." label.

    Returns (render_rows, net_profit). Render row kinds:
    header / account / total / subtotal.
    """
    settings = ReportSettings.get()
    detail = settings.pl_detail_rows or 10
    movements = _period_movements(from_date, to_date, ["revenue", "expense"])
    accounts = {a.id: a for a in ChartOfAccount.query.filter(
        ChartOfAccount.type.in_(["revenue", "expense", "contra-expense"])).all()}

    by_section = defaultdict(list)
    for aid, (dr, cr) in movements.items():
        a = accounts.get(aid)
        if a is None:
            continue
        contrib = cr - dr
        if dr == 0 and cr == 0:
            continue
        section = a.effective_pl_section() or (
            "other_income" if a.type == "revenue" else "other_operating")
        by_section[section].append({"code": a.code, "name": a.name, "contrib": contrib})

    rows, running = [], Decimal("0")
    for entry in settings.pl_structure():
        if "section" in entry:
            items = by_section.get(entry["section"], [])
            if not items:
                continue
            negate = bool(entry.get("negate"))
            disp = (lambda c: -c) if negate else (lambda c: c)
            total_contrib = sum(i["contrib"] for i in items)
            running += total_contrib
            items.sort(key=lambda i: -abs(i["contrib"]))
            shown, hidden = items[:detail], items[detail:]
            rows.append({"kind": "header", "label": entry["label"]})
            for i in shown:
                rows.append({"kind": "account", "code": i["code"], "name": i["name"],
                             "amount": float(disp(i["contrib"]))})
            if hidden:
                rows.append({"kind": "account", "code": "",
                             "name": f"Others ({len(hidden)} accounts)",
                             "amount": float(disp(sum(i["contrib"] for i in hidden)))})
            rows.append({"kind": "total", "label": f"Total {entry['label']}",
                         "amount": float(disp(total_contrib))})
        elif "subtotal" in entry:
            rows.append({"kind": "subtotal", "label": entry["label"],
                         "amount": float(running)})
    return rows, float(running)


def _build_excel_wb(title, headers, rows, col_widths=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center")

    hdr_row = 3
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=hdr_row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN

    for ri, row in enumerate(rows, hdr_row + 1):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = DATA_FONT
            c.border = THIN
            c.alignment = RIGHT if isinstance(val, (int, float, Decimal)) else LEFT_ALIGN

    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _build_pdf_landscape(title, headers, rows, col_widths=None):
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            rightMargin=10*mm, leftMargin=10*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Spacer(1, 6*mm))
    elements.append(Paragraph(f"Generated: {datetime.now():%Y-%m-%d %H:%M}",
                              styles["Normal"]))
    elements.append(Spacer(1, 4*mm))

    data = [headers] + [[str(c) if not isinstance(c, (int, float, Decimal)) else
                         f"{c:,.2f}" for c in row] for row in rows]
    if not col_widths:
        col_widths = [doc.width / len(headers)] * len(headers)
    else:
        pw = doc.width
        tw = sum(col_widths)
        col_widths = [pw * (w / tw) for w in col_widths]

    t = Table(data, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F7FB")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    t.setStyle(TableStyle(style))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════
# 1. DASHBOARD
# ═══════════════════════════════════════════════

@finance_bp.route("/")
@login_required
def dashboard():
    return render_template("finance/dashboard.html", now=datetime.utcnow())


# ═══════════════════════════════════════════════
# 2. GENERAL LEDGER
# ═══════════════════════════════════════════════

def _get_descendant_ids(account_id):
    ids = [account_id]
    for child in ChartOfAccount.query.filter_by(parent_id=account_id, is_active=True).all():
        ids.extend(_get_descendant_ids(child.id))
    return ids


def _get_leaf_descendant_ids(account_id):
    """Get only leaf (no-children) descendant IDs, excluding the head itself."""
    leaf_ids = []
    for child in ChartOfAccount.query.filter_by(parent_id=account_id, is_active=True).all():
        grand_children = ChartOfAccount.query.filter_by(parent_id=child.id, is_active=True).count()
        if grand_children == 0:
            leaf_ids.append(child.id)
        else:
            leaf_ids.extend(_get_leaf_descendant_ids(child.id))
    return list(set(leaf_ids))


def _get_ledger_sections(account_ids, from_date, to_date):
    """Per-account ledger: opening balance (all posted activity before the
    period), movements during the period with a running balance, and a
    closing balance labelled Dr/Cr."""
    sections = []
    for aid in account_ids:
        account = ChartOfAccount.query.get(aid)
        if not account:
            continue
        opening = Decimal("0")
        if from_date:
            odr, ocr = _get_account_balance(aid, from_date - timedelta(days=1))
            opening = odr - ocr
        q = JournalLine.query.join(JournalEntry).filter(
            JournalLine.account_id == aid,
            JournalEntry.is_posted == True,
        )
        if from_date:
            q = q.filter(JournalEntry.entry_date >= from_date)
        if to_date:
            q = q.filter(JournalEntry.entry_date <= to_date)
        q = q.order_by(JournalEntry.entry_date, JournalEntry.id)
        lines = q.all()
        rows = []
        balance = opening
        total_dr = total_cr = Decimal("0")
        for line in lines:
            dr = Decimal(str(line.debit))
            cr = Decimal(str(line.credit))
            balance += dr - cr
            total_dr += dr
            total_cr += cr
            rows.append({
                "date": line.entry.entry_date.strftime("%Y-%m-%d") if line.entry.entry_date else "",
                "voucher": line.entry.voucher_number or "",
                "description": line.description or line.entry.description or "",
                "debit": float(dr) if dr else 0,
                "credit": float(cr) if cr else 0,
                "balance": float(balance),
            })
        if not rows and opening == 0:
            # Nothing before or during the period — skip empty accounts when
            # rendering "all accounts" so the report stays readable.
            sections.append({"account": account, "rows": rows, "empty": True,
                             "opening": 0.0, "closing": 0.0,
                             "closing_side": "Dr", "opening_side": "Dr",
                             "total_debit": 0.0, "total_credit": 0.0,
                             "subtotal": 0.0})
            continue
        sections.append({
            "account": account,
            "rows": rows,
            "empty": False,
            "opening": float(abs(opening)),
            "opening_side": "Dr" if opening >= 0 else "Cr",
            "total_debit": float(total_dr),
            "total_credit": float(total_cr),
            "closing": float(abs(balance)),
            "closing_side": "Dr" if balance >= 0 else "Cr",
            "subtotal": float(balance),
        })
    return sections


@finance_bp.route("/ledger")
@login_required
def ledger():
    all_accounts = ChartOfAccount.query.filter_by(is_active=True).order_by(ChartOfAccount.code).all()
    child_ids = {r[0] for r in db.session.query(ChartOfAccount.parent_id).filter(
        ChartOfAccount.parent_id.isnot(None)).distinct().all()}
    heads = [a for a in all_accounts if a.parent_id is not None and a.id in child_ids]
    leaf_accounts = [a for a in all_accounts if a.id not in child_ids]

    from_date, to_date, periods, selected_period_id, filter_mode, from_str, to_str = _resolve_period()

    mode = request.args.get("mode", "all")
    selection_mode = request.args.get("selection_mode", "custom")
    account_ids_str = request.args.get("account_ids", "")
    head_ids_str = request.args.get("head_ids", "")

    resolved_ids = []
    if mode == "all":
        resolved_ids = [a.id for a in all_accounts]
    elif mode == "select":
        if selection_mode == "custom" and account_ids_str:
            resolved_ids = [int(x) for x in account_ids_str.split(",") if x.strip().isdigit()]
        elif selection_mode == "head" and head_ids_str:
            head_ids = [int(x) for x in head_ids_str.split(",") if x.strip().isdigit()]
            for hid in head_ids:
                resolved_ids.extend(_get_leaf_descendant_ids(hid))
            resolved_ids = list(set(resolved_ids))

    account_sections = _get_ledger_sections(resolved_ids, from_date, to_date) if resolved_ids else []
    account_sections = [s for s in account_sections if not s.get("empty")]

    fmt = request.args.get("format")
    if fmt and account_sections:
        headers = ["Date", "Voucher #", "Description", "Debit", "Credit", "Balance"]
        if fmt == "excel":
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            for sec in account_sections:
                data = ([["", "", "Opening Balance", "", "",
                          f"{sec['opening']:,.2f} {sec['opening_side']}"]] +
                        [[r["date"], r["voucher"], r["description"], r["debit"], r["credit"], r["balance"]]
                         for r in sec["rows"]])
                ws = wb.create_sheet(title=sec["account"].code[:31])
                ws.merge_cells("A1:F1")
                ws.cell(row=1, column=1, value=f"{sec['account'].code} - {sec['account'].name}").font = TITLE_FONT
                for ci, h in enumerate(headers, 1):
                    c = ws.cell(row=3, column=ci, value=h)
                    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN
                for ri, row in enumerate(data, 4):
                    for ci, val in enumerate(row, 1):
                        c = ws.cell(row=ri, column=ci, value=val)
                        c.font = DATA_FONT; c.border = THIN
                        c.alignment = RIGHT if isinstance(val, (int, float, Decimal)) else LEFT_ALIGN
                tr = 4 + len(data)
                ws.cell(row=tr, column=3, value="Period Movement").font = BOLD_FONT
                ws.cell(row=tr, column=4, value=sec["total_debit"]).font = BOLD_FONT
                ws.cell(row=tr, column=5, value=sec["total_credit"]).font = BOLD_FONT
                ws.cell(row=tr + 1, column=5, value="Closing Balance").font = BOLD_FONT
                ws.cell(row=tr + 1, column=6,
                        value=f"{sec['closing']:,.2f} {sec['closing_side']}").font = BOLD_FONT
            out = BytesIO(); wb.save(out); out.seek(0)
            return send_file(out, as_attachment=True, download_name="general_ledger.xlsx",
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        if fmt == "pdf":
            all_data = []
            for sec in account_sections:
                all_data.append([sec["account"].code, sec["account"].name, "", "", "", ""])
                all_data.append(["", "", "Opening Balance", "", "",
                                 f"{sec['opening']:,.2f} {sec['opening_side']}"])
                for r in sec["rows"]:
                    all_data.append([r["date"], r["voucher"], r["description"], r["debit"], r["credit"], r["balance"]])
                all_data.append(["", "", "Period Movement", sec["total_debit"], sec["total_credit"], ""])
                all_data.append(["", "", "", "", "Closing Balance",
                                 f"{sec['closing']:,.2f} {sec['closing_side']}"])
                all_data.append(["", "", "", "", "", ""])
            hdrs = ["Date", "Voucher #", "Description", "Debit", "Credit", "Balance"]
            pdf_out = _build_pdf_landscape("General Ledger", hdrs, all_data, [24, 28, 60, 24, 24, 24])
            return send_file(pdf_out, as_attachment=True, download_name="general_ledger.pdf",
                             mimetype="application/pdf")

    return render_template("finance/ledger.html",
                           account_sections=account_sections,
                           heads=heads, leaf_accounts=leaf_accounts,
                           mode=mode, selection_mode=selection_mode,
                           selected_account_ids=account_ids_str,
                           selected_head_ids=head_ids_str,
                           from_date=from_date, to_date=to_date,
                           periods=periods, selected_period_id=selected_period_id,
                           filter_mode=filter_mode, from_str=from_str, to_str=to_str,
                           now=datetime.utcnow())


# ═══════════════════════════════════════════════
# 3. TRIAL BALANCE
# ═══════════════════════════════════════════════

@finance_bp.route("/trial-balance")
@login_required
def trial_balance():
    from_date, to_date, periods, selected_period_id, filter_mode, from_str, to_str = _resolve_period()
    as_of = to_date or date.today()

    opening_as_of = None
    if from_date:
        opening_as_of = from_date - timedelta(days=1)

    opening_balances = _all_account_balances(opening_as_of) if opening_as_of else []
    closing_balances = _all_account_balances(as_of)

    # Index closing by account code
    closing_map = {}
    for b in closing_balances:
        closing_map[b.code] = b

    opening_map = {}
    for b in opening_balances:
        opening_map[b.code] = b

    all_codes = set(closing_map.keys()) | set(opening_map.keys())

    rows = []
    total_dr_op = Decimal("0")
    total_cr_op = Decimal("0")
    total_dr_mv = Decimal("0")
    total_cr_mv = Decimal("0")
    total_dr_cl = Decimal("0")
    total_cr_cl = Decimal("0")

    for code in sorted(all_codes):
        cb = closing_map.get(code)
        ob = opening_map.get(code)

        dr_op = ob.dr if ob else Decimal("0")
        cr_op = ob.cr if ob else Decimal("0")
        dr_cl = cb.dr if cb else Decimal("0")
        cr_cl = cb.cr if cb else Decimal("0")
        dr_mv = dr_cl - dr_op
        cr_mv = cr_cl - cr_op

        if dr_cl == cr_cl == 0 and dr_op == cr_op == 0:
            continue

        name = (cb or ob).name
        type_ = ACCOUNT_TYPES.get((cb or ob).type, (cb or ob).type)

        total_dr_op += dr_op
        total_cr_op += cr_op
        total_dr_mv += dr_mv
        total_cr_mv += cr_mv
        total_dr_cl += dr_cl
        total_cr_cl += cr_cl

        rows.append({
            "code": code,
            "name": name,
            "type": type_,
            "dr_opening": float(dr_op),
            "cr_opening": float(cr_op),
            "dr_movement": float(dr_mv),
            "cr_movement": float(cr_mv),
            "dr_closing": float(dr_cl),
            "cr_closing": float(cr_cl),
        })

    # Roll-up subtotals per account class (codes start with the class digit:
    # 1 Assets .. 5 Expenses), inserted after each class's rows.
    CLASS_NAMES = {"1": "Assets", "2": "Liabilities", "3": "Equity",
                   "4": "Revenue", "5": "Expenses"}
    grouped = []
    cls_tot = None
    prev_cls = None

    def close_class(g, tot, cls):
        if tot and cls in CLASS_NAMES:
            g.append({"code": "", "name": f"Total {CLASS_NAMES[cls]}", "type": "",
                      "is_subtotal": True, **{k: tot[k] for k in tot}})

    for r in rows:
        cls = (r["code"] or "?")[0]
        if cls != prev_cls:
            close_class(grouped, cls_tot, prev_cls)
            cls_tot = {k: 0.0 for k in ("dr_opening", "cr_opening", "dr_movement",
                                        "cr_movement", "dr_closing", "cr_closing")}
            prev_cls = cls
        grouped.append(r)
        for k in cls_tot:
            cls_tot[k] += r[k]
    close_class(grouped, cls_tot, prev_cls)
    rows = grouped

    fmt = request.args.get("format")
    headers = ["Code", "Account", "Type", "Dr Opening", "Cr Opening",
               "Dr Movement", "Cr Movement", "Dr Closing", "Cr Closing"]
    if fmt == "excel":
        data = [[r["code"], r["name"], r["type"],
                 r["dr_opening"], r["cr_opening"],
                 r["dr_movement"], r["cr_movement"],
                 r["dr_closing"], r["cr_closing"]] for r in rows]
        data.append(["", "TOTAL", "",
                      float(total_dr_op), float(total_cr_op),
                      float(total_dr_mv), float(total_cr_mv),
                      float(total_dr_cl), float(total_cr_cl)])
        wb_out = _build_excel_wb(f"Trial Balance as of {as_of}", headers, data,
                                 [10, 36, 14, 16, 16, 16, 16, 16, 16])
        return send_file(wb_out, as_attachment=True, download_name="trial_balance.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    if fmt == "pdf":
        data = [[r["code"], r["name"], r["type"],
                 f"{r['dr_opening']:,.2f}", f"{r['cr_opening']:,.2f}",
                 f"{r['dr_movement']:,.2f}", f"{r['cr_movement']:,.2f}",
                 f"{r['dr_closing']:,.2f}", f"{r['cr_closing']:,.2f}"] for r in rows]
        data.append(["", "TOTAL", "",
                      f"{float(total_dr_op):,.2f}", f"{float(total_cr_op):,.2f}",
                      f"{float(total_dr_mv):,.2f}", f"{float(total_cr_mv):,.2f}",
                      f"{float(total_dr_cl):,.2f}", f"{float(total_cr_cl):,.2f}"])
        pdf_out = _build_pdf_landscape(f"Trial Balance as of {as_of}", headers, data,
                                        [12, 48, 14, 28, 28, 28, 28, 28, 28])
        return send_file(pdf_out, as_attachment=True, download_name="trial_balance.pdf",
                         mimetype="application/pdf")

    return render_template("finance/trial_balance.html", rows=rows,
                           total_dr_opening=float(total_dr_op),
                           total_cr_opening=float(total_cr_op),
                           total_dr_movement=float(total_dr_mv),
                           total_cr_movement=float(total_cr_mv),
                           total_dr_closing=float(total_dr_cl),
                           total_cr_closing=float(total_cr_cl),
                           as_of=as_of, from_date=from_date,
                           periods=periods, selected_period_id=selected_period_id,
                           filter_mode=filter_mode, from_str=from_str, to_str=to_str,
                           now=datetime.utcnow())


# ═══════════════════════════════════════════════
# 4. PROFIT & LOSS
# ═══════════════════════════════════════════════

@finance_bp.route("/profit-loss")
@login_required
def profit_loss():
    from_date, to_date, periods, selected_period_id, filter_mode, from_str, to_str = _resolve_period()
    if not from_date: from_date = date(date.today().year, 1, 1)
    if not to_date: to_date = date.today()

    pl_rows, net_profit = _pl_rows(from_date, to_date)

    fmt = request.args.get("format")
    if fmt == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "P&L"
        ws.merge_cells("A1:C1")
        ws.cell(row=1, column=1, value=f"Profit & Loss ({from_date} to {to_date})").font = TITLE_FONT
        r = 3
        for row in pl_rows:
            if row["kind"] == "header":
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
                c = ws.cell(row=r, column=1, value=row["label"])
                c.font = Font(bold=True, size=12); c.fill = HEADER_FILL
                c.font = Font(bold=True, size=12, color="FFFFFF")
            elif row["kind"] == "account":
                ws.cell(row=r, column=1, value=row["code"]).font = DATA_FONT
                ws.cell(row=r, column=2, value=row["name"]).font = DATA_FONT
                c = ws.cell(row=r, column=3, value=row["amount"])
                c.font = DATA_FONT; c.alignment = RIGHT
            elif row["kind"] == "total":
                ws.cell(row=r, column=2, value=row["label"]).font = BOLD_FONT
                c = ws.cell(row=r, column=3, value=row["amount"])
                c.font = BOLD_FONT; c.alignment = RIGHT
            else:  # subtotal / profit line
                ws.cell(row=r, column=2, value=row["label"]).font = Font(bold=True, size=12, color="1F4E79")
                c = ws.cell(row=r, column=3, value=row["amount"])
                c.font = Font(bold=True, size=12, color="1F4E79"); c.alignment = RIGHT
                r += 1  # blank spacer after each profit line
            r += 1
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 44
        ws.column_dimensions["C"].width = 18
        out = BytesIO(); wb.save(out); out.seek(0)
        return send_file(out, as_attachment=True, download_name="profit_loss.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    if fmt == "pdf":
        headers = ["Code", "Account / Section", "Amount"]
        data = []
        for row in pl_rows:
            if row["kind"] == "header":
                data.append(["", row["label"].upper(), ""])
            elif row["kind"] == "account":
                data.append([row["code"], row["name"], f"{row['amount']:,.2f}"])
            elif row["kind"] == "total":
                data.append(["", row["label"], f"{row['amount']:,.2f}"])
            else:
                data.append(["", row["label"], f"{row['amount']:,.2f}"])
                data.append(["", "", ""])
        pdf_out = _build_pdf_landscape(f"Profit & Loss ({from_date} to {to_date})",
                                        headers, data, [24, 66, 26])
        return send_file(pdf_out, as_attachment=True, download_name="profit_loss.pdf",
                         mimetype="application/pdf")

    return render_template("finance/profit_loss.html", pl_rows=pl_rows,
                           net_profit=net_profit,
                           from_date=from_date, to_date=to_date,
                           periods=periods, selected_period_id=selected_period_id,
                           filter_mode=filter_mode, from_str=from_str, to_str=to_str,
                           now=datetime.utcnow())


# ═══════════════════════════════════════════════
# 5. BALANCE SHEET
# ═══════════════════════════════════════════════

@finance_bp.route("/balance-sheet")
@login_required
def balance_sheet():
    _, to_date, periods, selected_period_id, filter_mode, from_str, to_str = _resolve_period()
    as_of = to_date or date.today()
    as_of_end = as_of

    balances = _all_account_balances(as_of_end)
    ni = _net_income(as_of_end)

    assets, liabilities, equity = [], [], []
    total_assets = total_liabilities = total_equity = Decimal("0")

    for b in balances:
        if b.type == "asset":
            bal = b.dr - b.cr
            if bal != 0:
                assets.append({"code": b.code, "name": b.name, "amount": float(bal)})
                total_assets += bal
        elif b.type == "liability":
            bal = b.cr - b.dr
            if bal != 0:
                liabilities.append({"code": b.code, "name": b.name, "amount": float(bal)})
                total_liabilities += bal
        elif b.type == "equity":
            bal = b.cr - b.dr
            if bal != 0:
                equity.append({"code": b.code, "name": b.name, "amount": float(bal)})
                total_equity += bal

    if ni >= 0:
        equity.append({"code": "", "name": "Net Income (Current Period)", "amount": float(ni)})
    else:
        equity.append({"code": "", "name": "Net Loss (Current Period)", "amount": float(ni)})
    total_equity += Decimal(str(ni))

    fmt = request.args.get("format")
    if fmt == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"
        ws.merge_cells("A1:C1")
        ws.cell(row=1, column=1, value=f"Balance Sheet as of {as_of_end}").font = TITLE_FONT

        def write_section(ws, sr, section_title, items, total_label, total_val):
            ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=3)
            ws.cell(row=sr, column=1, value=section_title).font = Font(bold=True, size=12)
            hdr = sr + 1
            for ci, h in enumerate(["Code", "Account", "Amount"], 1):
                c = ws.cell(row=hdr, column=ci, value=h)
                c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN
            for ri, item in enumerate(items, hdr + 1):
                ws.cell(row=ri, column=1, value=item["code"]).font = DATA_FONT
                ws.cell(row=ri, column=1).border = THIN
                ws.cell(row=ri, column=2, value=item["name"]).font = DATA_FONT
                ws.cell(row=ri, column=2).border = THIN
                ws.cell(row=ri, column=3, value=item["amount"]).font = DATA_FONT
                ws.cell(row=ri, column=3).border = THIN
                ws.cell(row=ri, column=3).alignment = RIGHT
            tr = hdr + len(items) + 1
            ws.cell(row=tr, column=2, value=total_label).font = BOLD_FONT
            ws.cell(row=tr, column=2).border = THIN
            ws.cell(row=tr, column=3, value=float(total_val)).font = BOLD_FONT
            ws.cell(row=tr, column=3).border = THIN
            ws.cell(row=tr, column=3).alignment = RIGHT
            return tr + 2

        nr = write_section(ws, 3, "ASSETS", assets, "Total Assets", total_assets)
        nr = write_section(ws, nr, "LIABILITIES", liabilities, "Total Liabilities", total_liabilities)
        nr = write_section(ws, nr, "EQUITY", equity, "Total Equity", total_equity)

        ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=3)
        ws.cell(row=nr, column=1, value="LIABILITIES + EQUITY").font = Font(bold=True, size=12)
        ws.cell(row=nr, column=3, value=float(total_liabilities + total_equity)).font = Font(bold=True, size=12)
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 20
        out = BytesIO(); wb.save(out); out.seek(0)
        return send_file(out, as_attachment=True, download_name="balance_sheet.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    if fmt == "pdf":
        headers = ["Code", "Account", "Amount"]
        all_data = ([[a["code"], a["name"], f"{a['amount']:,.2f}"] for a in assets] +
                    [["", "", ""]] +
                    [[l["code"], l["name"], f"{l['amount']:,.2f}"] for l in liabilities] +
                    [["", "", ""]] +
                    [[e["code"], e["name"], f"{e['amount']:,.2f}"] for e in equity])
        pdf_out = _build_pdf_landscape(f"Balance Sheet as of {as_of_end}", headers, all_data, [20, 60, 30])
        return send_file(pdf_out, as_attachment=True, download_name="balance_sheet.pdf",
                         mimetype="application/pdf")

    return render_template("finance/balance_sheet.html", assets=assets,
                           liabilities=liabilities, equity=equity,
                           total_assets=float(total_assets),
                           total_liabilities=float(total_liabilities),
                           total_equity=float(total_equity),
                           as_of=as_of_end,
                           periods=periods, selected_period_id=selected_period_id,
                           filter_mode=filter_mode, from_str=from_str, to_str=to_str,
                           now=datetime.utcnow())


# ═══════════════════════════════════════════════
# 6. SOCIE
# ═══════════════════════════════════════════════

@finance_bp.route("/socie")
@login_required
def socie():
    from_date, to_date, periods, selected_period_id, filter_mode, from_str, to_str = _resolve_period()
    if not from_date: from_date = date(date.today().year, 1, 1)
    if not to_date: to_date = date.today()

    ni = _net_income(to_date)
    eq_balances = _all_account_balances(to_date, ["equity"])

    opening_total = Decimal("0")
    movement_total = Decimal("0")
    rows = []

    for b in eq_balances:
        bal = b.cr - b.dr
        rows.append({"name": b.name, "opening": float(bal), "movement": 0, "closing": float(bal)})
        opening_total += bal
        closing_total = opening_total
    rows.append({"name": "Net Income / (Loss)", "opening": 0,
                 "movement": float(ni), "closing": float(ni)})
    movement_total += Decimal(str(ni))
    closing_total = opening_total + movement_total

    fmt = request.args.get("format")
    if fmt == "excel":
        headers = ["Component", "Opening Balance", "Movement", "Closing Balance"]
        data = [[r["name"], r["opening"], r["movement"], r["closing"]] for r in rows]
        data.append(["TOTAL", float(opening_total), float(movement_total), float(closing_total)])
        wb_out = _build_excel_wb(f"SOCIE ({from_date} to {to_date})", headers, data, [40, 20, 20, 20])
        return send_file(wb_out, as_attachment=True, download_name="socie.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    if fmt == "pdf":
        headers = ["Component", "Opening", "Movement", "Closing"]
        data = [[r["name"], f"{r['opening']:,.2f}", f"{r['movement']:,.2f}",
                 f"{r['closing']:,.2f}"] for r in rows]
        data.append(["TOTAL", f"{float(opening_total):,.2f}", f"{float(movement_total):,.2f}",
                     f"{float(closing_total):,.2f}"])
        pdf_out = _build_pdf_landscape(f"SOCIE ({from_date} to {to_date})", headers, data, [60, 30, 30, 30])
        return send_file(pdf_out, as_attachment=True, download_name="socie.pdf",
                         mimetype="application/pdf")

    return render_template("finance/socie.html", rows=rows,
                           opening_total=float(opening_total),
                           movement_total=float(movement_total),
                           closing_total=float(closing_total),
                           from_date=from_date, to_date=to_date,
                           periods=periods, selected_period_id=selected_period_id,
                           filter_mode=filter_mode, from_str=from_str, to_str=to_str,
                           now=datetime.utcnow())


# ═══════════════════════════════════════════════
# 7. CASH FLOW STATEMENT (Indirect Method)
# ═══════════════════════════════════════════════

@finance_bp.route("/cash-flow")
@login_required
def cash_flow():
    """Indirect-method cash flow driven by per-account cash_flow_activity tags.

    Net profit for the period, adjusted by the period's balance-sheet
    movements grouped by each account's effective activity tag (operating /
    investing / financing). Accounts tagged "cash" are the statement's
    subject: their movement is the target the three activity totals must
    reconcile to.
    """
    from_date, to_date, periods, selected_period_id, filter_mode, from_str, to_str = _resolve_period()
    if not from_date: from_date = date(date.today().year, 1, 1)
    if not to_date: to_date = date.today()
    opening_cutoff = from_date - timedelta(days=1)

    # Net profit for the period: sum of (cr - dr) over P&L accounts.
    pl_moves = _period_movements(from_date, to_date, ["revenue", "expense", "contra-expense"])
    net_profit = float(sum(cr - dr for dr, cr in pl_moves.values()))

    # Balance-sheet movements for the period, grouped by activity tag, and
    # within each activity by the account's level-3 head for readability.
    bs_moves = _period_movements(from_date, to_date, ["asset", "liability", "equity"])
    accounts = {a.id: a for a in ChartOfAccount.query.all()}

    def l3_head(acct):
        a = acct
        while a is not None and a.level > 3:
            a = a.parent
        return a.name if a is not None else acct.name

    groups = {"operating": defaultdict(float), "investing": defaultdict(float),
              "financing": defaultdict(float)}
    cash_movement = 0.0
    for aid, (dr, cr) in bs_moves.items():
        acct = accounts.get(aid)
        if acct is None:
            continue
        activity = acct.effective_cash_flow_activity() or "operating"
        if activity == "cash":
            cash_movement += float(dr - cr)
            continue
        # Cash effect of a balance-sheet movement: an asset build-up consumes
        # cash (-(dr-cr)); a liability/equity build-up provides it (+(cr-dr)).
        # Both reduce to (cr - dr) regardless of account type.
        effect = float(cr - dr)
        if effect:
            groups[activity][l3_head(acct)] += effect

    op_items = [("Net Profit / (Loss) for the period", net_profit)]
    op_items += [(f"(Increase) / Decrease in {name}" if v < 0 else
                  f"Decrease / (Increase) in {name}", v)
                 for name, v in sorted(groups["operating"].items())]
    inv_items = [(f"Movement in {name}", v) for name, v in sorted(groups["investing"].items())]
    fin_items = [(f"Movement in {name}", v) for name, v in sorted(groups["financing"].items())]

    net_operating = sum(v for _, v in op_items)
    net_investing = sum(v for _, v in inv_items)
    net_financing = sum(v for _, v in fin_items)
    net_change = net_operating + net_investing + net_financing

    # Reconciliation against the actual cash-account balances.
    cash_ids = [a.id for a in accounts.values()
                if (a.effective_cash_flow_activity() or "") == "cash" and a.level >= 5]

    def cash_balance(as_of):
        total = Decimal("0")
        for cid in cash_ids:
            dr, cr = _get_account_balance(cid, as_of)
            total += dr - cr
        return float(total)
    opening_cash = cash_balance(opening_cutoff)
    closing_cash = cash_balance(to_date)

    fmt = request.args.get("format")
    if fmt == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cash Flow"
        ws.merge_cells("A1:B1")
        ws.cell(row=1, column=1, value=f"Cash Flow Statement ({from_date} to {to_date})").font = TITLE_FONT

        def write_section(sr, title, items, total_label, total_val):
            ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=2)
            ws.cell(row=sr, column=1, value=title).font = Font(bold=True, size=12)
            r = sr + 1
            for name, val in items:
                ws.cell(row=r, column=1, value=name).font = DATA_FONT
                cc = ws.cell(row=r, column=2, value=val)
                cc.font = DATA_FONT; cc.alignment = RIGHT
                r += 1
            ws.cell(row=r, column=1, value=total_label).font = BOLD_FONT
            cc = ws.cell(row=r, column=2, value=total_val)
            cc.font = BOLD_FONT; cc.alignment = RIGHT
            return r + 2

        nr = write_section(3, "OPERATING ACTIVITIES", op_items,
                           "Net cash from operating activities", net_operating)
        nr = write_section(nr, "INVESTING ACTIVITIES", inv_items,
                           "Net cash from investing activities", net_investing)
        nr = write_section(nr, "FINANCING ACTIVITIES", fin_items,
                           "Net cash from financing activities", net_financing)
        for label, val in [("NET CHANGE IN CASH", net_change),
                           ("Opening cash & equivalents", opening_cash),
                           ("Closing cash & equivalents", closing_cash)]:
            ws.cell(row=nr, column=1, value=label).font = BOLD_FONT
            cc = ws.cell(row=nr, column=2, value=val)
            cc.font = BOLD_FONT; cc.alignment = RIGHT
            nr += 1
        ws.column_dimensions["A"].width = 46
        ws.column_dimensions["B"].width = 20
        out = BytesIO(); wb.save(out); out.seek(0)
        return send_file(out, as_attachment=True, download_name="cash_flow.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    if fmt == "pdf":
        headers = ["Item", "Amount"]
        pdf_data = ([["OPERATING ACTIVITIES", ""]] +
                    [[n, f"{v:,.2f}"] for n, v in op_items] +
                    [["Net cash from operating activities", f"{net_operating:,.2f}"], ["", ""]] +
                    [["INVESTING ACTIVITIES", ""]] +
                    [[n, f"{v:,.2f}"] for n, v in inv_items] +
                    [["Net cash from investing activities", f"{net_investing:,.2f}"], ["", ""]] +
                    [["FINANCING ACTIVITIES", ""]] +
                    [[n, f"{v:,.2f}"] for n, v in fin_items] +
                    [["Net cash from financing activities", f"{net_financing:,.2f}"], ["", ""]] +
                    [["NET CHANGE IN CASH", f"{net_change:,.2f}"],
                     ["Opening cash & equivalents", f"{opening_cash:,.2f}"],
                     ["Closing cash & equivalents", f"{closing_cash:,.2f}"]])
        pdf_out = _build_pdf_landscape(f"Cash Flow Statement ({from_date} to {to_date})",
                                       headers, pdf_data, [80, 30])
        return send_file(pdf_out, as_attachment=True, download_name="cash_flow.pdf",
                         mimetype="application/pdf")

    return render_template("finance/cash_flow.html", op_items=op_items,
                           inv_items=inv_items, fin_items=fin_items,
                           net_operating=net_operating, net_investing=net_investing,
                           net_financing=net_financing, net_change=net_change,
                           opening_cash=opening_cash, closing_cash=closing_cash,
                           cash_movement=cash_movement,
                           from_date=from_date, to_date=to_date,
                           periods=periods, selected_period_id=selected_period_id,
                           filter_mode=filter_mode, from_str=from_str, to_str=to_str,
                           now=datetime.utcnow())

