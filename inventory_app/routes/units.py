import json
from io import BytesIO
from flask import Blueprint, render_template, redirect, url_for, flash, request, session, send_file, jsonify
from flask_login import login_required
from sqlalchemy import or_
from ..extensions import db
from ..models.unit import InvUnit

inv_units_bp = Blueprint("inv_units", __name__, url_prefix="/inventory/units")


@inv_units_bp.route("/check")
@login_required
def check_duplicate():
    field = request.args.get("field")
    value = request.args.get("value", "").strip()
    exclude = request.args.get("exclude", type=int)
    if field not in ("name", "abbreviation") or not value:
        return {"exists": False}
    q = InvUnit.query.filter(getattr(InvUnit, field) == value)
    if exclude:
        q = q.filter(InvUnit.id != exclude)
    exists = q.first() is not None
    return {"exists": exists}


@inv_units_bp.route("/template")
@login_required
def download_template():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Units"
    ws.append(["Name", "Abbreviation", "Explanation"])
    ws.append(["Kilogram", "kg", "Weight in kilograms"])
    ws.append(["Meter", "m", "Length in meters"])
    ws.append(["Piece", "pcs", "Individual unit count"])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name="unit_import_template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@inv_units_bp.route("/")
@login_required
def list_units():
    units = InvUnit.query.order_by(InvUnit.name).all()
    return render_template("units/list_inv.html", units=units)


@inv_units_bp.route("/create", methods=["GET", "POST"])
@login_required
def create_unit():
    if request.method == "POST":
        name = request.form["name"].strip()
        abbreviation = request.form["abbreviation"].strip()
        explanation = request.form.get("explanation", "").strip()

        if InvUnit.query.filter_by(name=name).first():
            flash(f"Unit '{name}' already exists.", "error")
            return render_template("units/form_inv.html", unit=None, name=name, abbreviation=abbreviation, explanation=explanation)

        if InvUnit.query.filter_by(abbreviation=abbreviation).first():
            flash(f"Abbreviation '{abbreviation}' already exists.", "error")
            return render_template("units/form_inv.html", unit=None, name=name, abbreviation=abbreviation, explanation=explanation)

        unit = InvUnit(name=name, abbreviation=abbreviation, explanation=explanation or None)
        db.session.add(unit)
        db.session.commit()
        flash(f"Unit '{name}' created.", "success")
        return redirect(url_for("inv_units.list_units"))

    return render_template("units/form_inv.html", unit=None, name="", abbreviation="", explanation="")


@inv_units_bp.route("/edit/<int:id>", methods=["GET", "POST"])
@login_required
def edit_unit(id):
    unit = InvUnit.query.get_or_404(id)
    if request.method == "POST":
        name = request.form["name"].strip()
        abbreviation = request.form["abbreviation"].strip()
        explanation = request.form.get("explanation", "").strip()

        dup = InvUnit.query.filter(InvUnit.id != id, InvUnit.name == name).first()
        if dup:
            flash(f"Unit '{name}' already exists.", "error")
            return render_template("units/form_inv.html", unit=unit)

        dup = InvUnit.query.filter(InvUnit.id != id, InvUnit.abbreviation == abbreviation).first()
        if dup:
            flash(f"Abbreviation '{abbreviation}' already exists.", "error")
            return render_template("units/form_inv.html", unit=unit)

        unit.name = name
        unit.abbreviation = abbreviation
        unit.explanation = explanation or None
        unit.is_active = request.form.get("is_active") == "on"
        db.session.commit()
        flash(f"Unit '{name}' updated.", "success")
        return redirect(url_for("inv_units.list_units"))

    return render_template("units/form_inv.html", unit=unit)


@inv_units_bp.route("/delete/<int:id>")
@login_required
def delete_unit(id):
    unit = InvUnit.query.get_or_404(id)
    db.session.delete(unit)
    db.session.commit()
    flash(f"Unit '{unit.name}' deleted.", "success")
    return redirect(url_for("inv_units.list_units"))


@inv_units_bp.route("/quick-create", methods=["POST"])
@login_required
def quick_create():
    name = request.form.get("name", "").strip()
    abbrev = request.form.get("abbreviation", "").strip()
    expl = request.form.get("explanation", "").strip()

    if not name or not abbrev:
        return jsonify(ok=False, error="Name and abbreviation are required.")

    if InvUnit.query.filter_by(name=name).first():
        return jsonify(ok=False, error=f"Unit '{name}' already exists.")

    if InvUnit.query.filter_by(abbreviation=abbrev).first():
        return jsonify(ok=False, error=f"Abbreviation '{abbrev}' already exists.")

    unit = InvUnit(name=name, abbreviation=abbrev, explanation=expl or None)
    db.session.add(unit)
    db.session.commit()
    return jsonify(ok=True, id=unit.id, name=unit.name, abbreviation=unit.abbreviation)


@inv_units_bp.route("/upload-excel", methods=["POST"])
@login_required
def upload_excel():
    file = request.files.get("file")
    if not file:
        flash("No file selected.", "error")
        return redirect(url_for("inv_units.list_units"))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
        ws = wb.active
        if not ws:
            flash("Excel file is empty.", "error")
            return redirect(url_for("inv_units.list_units"))

        rows = []
        errors = []
        seen_names = set()
        seen_abbrevs = set()

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            name = (str(row[0] or "").strip())
            abbrev = (str(row[1] or "").strip())
            expl = (str(row[2] or "").strip()) if len(row) > 2 else ""

            if not name and not abbrev:
                continue

            if not name:
                errors.append(f"Row {i}: Name is required.")
                continue
            if not abbrev:
                errors.append(f"Row {i}: Abbreviation is required.")
                continue

            if name in seen_names:
                errors.append(f"Row {i}: Duplicate name '{name}' in spreadsheet.")
                continue
            if abbrev in seen_abbrevs:
                errors.append(f"Row {i}: Duplicate abbreviation '{abbrev}' in spreadsheet.")
                continue

            existing = InvUnit.query.filter(or_(InvUnit.name == name, InvUnit.abbreviation == abbrev)).first()
            if existing:
                errors.append(f"Row {i}: '{name}' ({abbrev}) already exists in database.")
                continue

            seen_names.add(name)
            seen_abbrevs.add(abbrev)
            rows.append({"name": name, "abbreviation": abbrev, "explanation": expl})

        if not rows:
            flash("No valid rows found to import." + (" " + "; ".join(errors) if errors else ""), "error")
            return redirect(url_for("inv_units.list_units"))

        session["batch_units"] = rows
        flash(f"Parsed {len(rows)} unit(s) from Excel." + (" Warnings: " + "; ".join(errors) if errors else ""), "success" if not errors else "warning")
        return redirect(url_for("inv_units.batch_editor"))

    except Exception as e:
        flash(f"Failed to parse Excel file: {e}", "error")
        return redirect(url_for("inv_units.list_units"))


@inv_units_bp.route("/batch", methods=["GET", "POST"])
@login_required
def batch_editor():
    if request.method == "POST":
        raw = request.form.get("data", "[]")
        try:
            rows = json.loads(raw)
        except (json.JSONDecodeError, TypeError):
            flash("Invalid data submitted.", "error")
            return redirect(url_for("inv_units.list_units"))

        created = 0
        errors = []
        for r in rows:
            name = r.get("name", "").strip()
            abbrev = r.get("abbreviation", "").strip()
            expl = r.get("explanation", "").strip()
            row_marked = r.get("_delete", False)

            if row_marked:
                continue
            if not name or not abbrev:
                errors.append(f"'{name or '(empty)'}' — name and abbreviation are required.")
                continue

            if InvUnit.query.filter_by(name=name).first():
                errors.append(f"'{name}' — already exists.")
                continue
            if InvUnit.query.filter_by(abbreviation=abbrev).first():
                errors.append(f"'{abbrev}' — abbreviation already exists.")
                continue

            unit = InvUnit(name=name, abbreviation=abbrev, explanation=expl or None)
            db.session.add(unit)
            created += 1

        db.session.commit()
        session.pop("batch_units", None)
        msg = f"{created} unit(s) created."
        if errors:
            msg += " Errors: " + "; ".join(errors)
        flash(msg, "success" if not errors else "warning")
        return redirect(url_for("inv_units.list_units"))

    rows = session.pop("batch_units", [])
    return render_template("units/batch_editor.html", rows=rows)
