import json
from io import BytesIO
from flask import Blueprint, render_template, redirect, url_for, flash, request, session, send_file
from flask_login import login_required, current_user
from sqlalchemy import or_
from ..extensions import db
from ..models.product import InvProduct
from ..models.category import InvCategory
from ..models.unit import InvUnit
from ..models.stock_movement import InvStockMovement
from shared.ledger_utils import create_entity_account
from shared.permissions import deny_page

inv_prod_bp = Blueprint("inv_products", __name__, url_prefix="/inventory/products")


@inv_prod_bp.route("/")
@login_required
def list_products():
    q = request.args.get("q", "")
    cat_id = request.args.get("category_id", type=int)
    query = InvProduct.query
    if q:
        query = query.filter(InvProduct.name.ilike(f"%{q}%") | InvProduct.sku.ilike(f"%{q}%"))
    if cat_id:
        query = query.filter_by(category_id=cat_id)
    products = query.order_by(InvProduct.name).all()
    categories = InvCategory.query.filter_by(is_active=True).all()
    return render_template("products/list_inv.html", products=products, categories=categories)


@inv_prod_bp.route("/create", methods=["GET", "POST"])
@login_required
def create_product():
    if deny_page("products", "create"):
        return redirect(url_for("inv_products.list_products"))
    if request.method == "POST":
        prod = InvProduct(
            sku=request.form["sku"],
            name=request.form["name"],
            description=request.form.get("description", ""),
            category_id=request.form.get("category_id", type=int) or None,
            unit_price=request.form.get("unit_price", 0, type=float),
            cost_price=request.form.get("cost_price", 0, type=float),
            reorder_level=request.form.get("reorder_level", 0, type=int),
            current_stock=request.form.get("current_stock", 0, type=int),
            unit=request.form.get("unit", "pcs"),
        )
        db.session.add(prod)
        db.session.flush()
        create_entity_account("product", prod.id, f"{prod.name} ({prod.sku})")
        db.session.commit()
        flash(f"Product created — ledger account '{prod.name} ({prod.sku})' added under Inventory", "success")
        return redirect(url_for("inv_products.list_products"))
    categories = InvCategory.query.filter_by(is_active=True).all()
    units = InvUnit.query.filter_by(is_active=True).order_by(InvUnit.name).all()
    return render_template("products/form_inv.html", product=None, categories=categories, units=units)


@inv_prod_bp.route("/edit/<int:id>", methods=["GET", "POST"])
@login_required
def edit_product(id):
    if deny_page("products", "edit"):
        return redirect(url_for("inv_products.list_products"))
    prod = InvProduct.query.get_or_404(id)
    if request.method == "POST":
        prod.sku = request.form["sku"]
        prod.name = request.form["name"]
        prod.description = request.form.get("description", "")
        prod.category_id = request.form.get("category_id", type=int) or None
        prod.unit_price = request.form.get("unit_price", 0, type=float)
        prod.cost_price = request.form.get("cost_price", 0, type=float)
        prod.reorder_level = request.form.get("reorder_level", 0, type=int)
        prod.unit = request.form.get("unit", "pcs")
        prod.is_active = request.form.get("is_active") == "on"
        create_entity_account("product", prod.id, f"{prod.name} ({prod.sku})")
        db.session.commit()
        flash("Product updated", "success")
        return redirect(url_for("inv_products.list_products"))
    categories = InvCategory.query.filter_by(is_active=True).all()
    units = InvUnit.query.filter_by(is_active=True).order_by(InvUnit.name).all()
    return render_template("products/form_inv.html", product=prod, categories=categories, units=units)


@inv_prod_bp.route("/delete/<int:id>")
@login_required
def delete_product(id):
    if deny_page("products", "delete"):
        return redirect(url_for("inv_products.list_products"))
    prod = InvProduct.query.get_or_404(id)
    if prod.po_items.count() > 0 or prod.so_items.count() > 0:
        flash("Cannot delete product with order history", "error")
    else:
        db.session.delete(prod)
        db.session.commit()
        flash("Product deleted", "success")
    return redirect(url_for("inv_products.list_products"))


@inv_prod_bp.route("/adjust-stock/<int:id>", methods=["GET", "POST"])
@login_required
def adjust_stock(id):
    prod = InvProduct.query.get_or_404(id)
    if request.method == "POST":
        qty = request.form.get("quantity", 0, type=int)
        note = request.form.get("notes", "")
        if qty == 0:
            flash("Quantity must be non-zero", "error")
        else:
            mtype = "adjustment_in" if qty > 0 else "adjustment_out"
            prod.current_stock += qty
            InvStockMovement(
                product_id=prod.id, type=mtype, quantity=abs(qty),
                notes=note or f"Manual adjustment of {qty}",
                created_by=current_user.id  # noqa
            )
            db.session.add(prod)
            db.session.commit()
            flash(f"Stock adjusted by {qty}. New stock: {prod.current_stock}", "success")
            return redirect(url_for("inv_products.list_products"))
    return render_template("products/adjust_stock_inv.html", product=prod)


@inv_prod_bp.route("/check")
@login_required
def check_duplicate():
    field = request.args.get("field")
    value = request.args.get("value", "").strip()
    exclude = request.args.get("exclude", type=int)
    if field not in ("name", "sku") or not value:
        return {"exists": False}
    q = InvProduct.query.filter(getattr(InvProduct, field) == value)
    if exclude:
        q = q.filter(InvProduct.id != exclude)
    exists = q.first() is not None
    return {"exists": exists}


@inv_prod_bp.route("/template")
@login_required
def download_template():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["SKU", "Name", "Description", "Category", "Unit Price", "Cost Price", "Reorder Level", "Current Stock", "Unit"])
    ws.append(["SKU-001", "Sample Product A", "Description for product A", "Category Name", "100", "70", "10", "50", "pcs"])
    ws.append(["SKU-002", "Sample Product B", "Description for product B", "Category Name", "250", "180", "5", "20", "kg"])
    ws.append(["SKU-003", "Sample Product C", "Description for product C", "Category Name", "50", "30", "20", "100", "m"])
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 10
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name="product_import_template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@inv_prod_bp.route("/upload-excel", methods=["POST"])
@login_required
def upload_excel():
    file = request.files.get("file")
    if not file:
        flash("No file selected.", "error")
        return redirect(url_for("inv_products.list_products"))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
        ws = wb.active
        if not ws:
            flash("Excel file is empty.", "error")
            return redirect(url_for("inv_products.list_products"))

        rows = []
        errors = []
        seen_skus = set()
        seen_names = set()

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            sku = (str(row[0] or "").strip()).upper() if row[0] is not None else ""
            name = (str(row[1] or "").strip()) if len(row) > 1 and row[1] is not None else ""
            desc = (str(row[2] or "").strip()) if len(row) > 2 and row[2] is not None else ""
            cat_name = (str(row[3] or "").strip()) if len(row) > 3 and row[3] is not None else ""
            unit_price = 0.0
            cost_price = 0.0
            reorder = 0
            stock = 0
            unit = "pcs"

            try:
                if len(row) > 4 and row[4] is not None:
                    unit_price = float(str(row[4]).replace(",", ""))
            except (ValueError, TypeError):
                errors.append(f"Row {i}: Invalid Unit Price.")

            try:
                if len(row) > 5 and row[5] is not None:
                    cost_price = float(str(row[5]).replace(",", ""))
            except (ValueError, TypeError):
                errors.append(f"Row {i}: Invalid Cost Price.")

            try:
                if len(row) > 6 and row[6] is not None:
                    reorder = int(float(str(row[6]).replace(",", "")))
            except (ValueError, TypeError):
                errors.append(f"Row {i}: Invalid Reorder Level.")

            try:
                if len(row) > 7 and row[7] is not None:
                    stock = int(float(str(row[7]).replace(",", "")))
            except (ValueError, TypeError):
                errors.append(f"Row {i}: Invalid Current Stock.")

            if len(row) > 8 and row[8] is not None:
                unit = (str(row[8]).strip())

            if not sku and not name:
                continue
            if not sku:
                errors.append(f"Row {i}: SKU is required.")
                continue
            if not name:
                errors.append(f"Row {i}: Name is required.")
                continue

            if sku in seen_skus:
                errors.append(f"Row {i}: Duplicate SKU '{sku}' in spreadsheet.")
                continue
            if name in seen_names:
                errors.append(f"Row {i}: Duplicate name '{name}' in spreadsheet.")
                continue

            existing = InvProduct.query.filter(
                or_(InvProduct.sku == sku, InvProduct.name == name)
            ).first()
            if existing:
                errors.append(f"Row {i}: '{name}' ({sku}) already exists in database.")
                continue

            seen_skus.add(sku)
            seen_names.add(name)
            rows.append({
                "sku": sku,
                "name": name,
                "description": desc,
                "category_name": cat_name,
                "unit_price": unit_price,
                "cost_price": cost_price,
                "reorder_level": reorder,
                "current_stock": stock,
                "unit": unit,
            })

        if not rows:
            flash("No valid rows found to import." + (" " + "; ".join(errors) if errors else ""), "error")
            return redirect(url_for("inv_products.list_products"))

        session["batch_products"] = rows
        flash(f"Parsed {len(rows)} product(s) from Excel." + (" Warnings: " + "; ".join(errors) if errors else ""), "success" if not errors else "warning")
        return redirect(url_for("inv_products.batch_editor"))

    except Exception as e:
        flash(f"Failed to parse Excel file: {e}", "error")
        return redirect(url_for("inv_products.list_products"))


@inv_prod_bp.route("/batch", methods=["GET", "POST"])
@login_required
def batch_editor():
    if request.method == "POST":
        raw = request.form.get("data", "[]")
        try:
            rows = json.loads(raw)
        except (json.JSONDecodeError, TypeError):
            flash("Invalid data submitted.", "error")
            return redirect(url_for("inv_products.list_products"))

        created = 0
        errors = []
        for r in rows:
            sku = r.get("sku", "").strip().upper()
            name = r.get("name", "").strip()
            desc = r.get("description", "").strip()
            cat_name = r.get("category_name", "").strip()
            unit_price = r.get("unit_price", 0)
            cost_price = r.get("cost_price", 0)
            reorder = r.get("reorder_level", 0)
            stock = r.get("current_stock", 0)
            unit = r.get("unit", "pcs").strip()
            row_marked = r.get("_delete", False)

            if row_marked:
                continue
            if not sku or not name:
                errors.append(f"'{name or '(empty)'}' — SKU and name are required.")
                continue

            if InvProduct.query.filter_by(sku=sku).first():
                errors.append(f"'{sku}' — SKU already exists.")
                continue
            if InvProduct.query.filter_by(name=name).first():
                errors.append(f"'{name}' — name already exists.")
                continue

            category_id = None
            if cat_name:
                cat = InvCategory.query.filter_by(name=cat_name).first()
                if cat:
                    category_id = cat.id
                else:
                    errors.append(f"'{name}' — category '{cat_name}' not found.")
                    continue

            try:
                unit_price = float(unit_price)
            except (ValueError, TypeError):
                unit_price = 0
            try:
                cost_price = float(cost_price)
            except (ValueError, TypeError):
                cost_price = 0
            try:
                reorder = int(reorder)
            except (ValueError, TypeError):
                reorder = 0
            try:
                stock = int(stock)
            except (ValueError, TypeError):
                stock = 0

            prod = InvProduct(
                sku=sku,
                name=name,
                description=desc or None,
                category_id=category_id,
                unit_price=unit_price,
                cost_price=cost_price,
                reorder_level=reorder,
                current_stock=stock,
                unit=unit,
            )
            db.session.add(prod)
            db.session.flush()
            create_entity_account("product", prod.id, f"{prod.name} ({prod.sku})")
            created += 1

        db.session.commit()
        session.pop("batch_products", None)
        msg = f"{created} product(s) created."
        if errors:
            msg += " Errors: " + "; ".join(errors)
        flash(msg, "success" if not errors else "warning")
        return redirect(url_for("inv_products.list_products"))

    rows = session.pop("batch_products", [])
    units = InvUnit.query.filter_by(is_active=True).order_by(InvUnit.name).all()
    return render_template("products/batch_editor.html", rows=rows, units=units)
