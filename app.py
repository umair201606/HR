import os
import io
import re
import sys
import zipfile
from datetime import datetime
from collections import OrderedDict

from flask import (
    Flask, render_template, request, jsonify, send_file, session
)
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import (
    Table, TableStyle, Paragraph, Spacer, PageBreak, FrameBreak
)
from reportlab.platypus.doctemplate import BaseDocTemplate, Frame, PageTemplate

# PyInstaller support: when bundled, files are in sys._MEIPASS
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__,
            template_folder=os.path.join(BASE_DIR, "templates"),
            static_folder=os.path.join(BASE_DIR, "static"))
app.secret_key = os.urandom(24)
app.config["UPLOAD_FOLDER"] = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "uploads")
app.config["OUTPUT_FOLDER"] = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "output")
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024

os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
os.makedirs(app.config["OUTPUT_FOLDER"], exist_ok=True)

COMPANY_NAME = "Solarkon (Private) Limited"
SALARY_COLUMNS = [
    "Month", "Name", "Designation", "Basic Salary",
    "Medical Allowance", "House Rent", "Fuel", "HP", "EB",
    "PF D/LE", "GT", "Taxable Salary",
    "EPFC", "ADV.", "DED.", "WHT", "EOBI", "TD", "Payable", "Target Salary"
]
EMPLOYEE_COLUMNS = [
    "CNIC", "Phone", "Email", "Bank Name", "Account Title", "Account Number"
]



def parse_currency(val):
    if val is None:
        return 0
    val = str(val).strip()
    if val in ("", "-", "--", "N/A", "n/a"):
        return 0
    val = val.replace(",", "").replace(" ", "").replace('"', "")
    try:
        return float(val)
    except ValueError:
        return 0


def format_currency(val):
    if val is None or val == 0:
        return "-"
    if val == int(val):
        return f"{int(val):,}"
    return f"{val:,.2f}"


def parse_month_key(month_str):
    month_map = {
        "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4,
        "May": 5, "Jun": 6, "Jul": 7, "Aug": 8,
        "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
    }
    m = re.match(r"(\w+)-(\d{4})", month_str.strip())
    if m:
        mon, year = m.groups()
        return (int(year), month_map.get(mon[:3], 0), month_str)
    return (0, 0, month_str)


def format_month_display(month_str):
    m = {
        "Jan": "January", "Feb": "February", "Mar": "March", "Apr": "April",
        "May": "May", "Jun": "June", "Jul": "July", "Aug": "August",
        "Sep": "September", "Oct": "October", "Nov": "November", "Dec": "December"
    }
    match = re.match(r"(\w+)-(\d{4})", month_str.strip())
    if match:
        full = m.get(match.group(1)[:3], match.group(1))
        return f"{full} {match.group(2)}"
    return month_str


def read_csv_robust(filepath):
    for enc in ["utf-8-sig", "utf-8", "latin-1", "cp1252"]:
        try:
            with open(filepath, "r", encoding=enc) as f:
                content = f.read()
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    else:
        raise ValueError("Could not read CSV with any supported encoding")

    content = content.replace("\r\n", "\n").replace("\r", "\n")
    lines = content.strip().split("\n")
    if not lines:
        raise ValueError("CSV file is empty")

    header = [h.strip() for h in lines[0].split(",")]

    def parse_line(line):
        vals, cur, in_q = [], "", False
        for ch in line:
            if ch == '"':
                in_q = not in_q
            elif ch == "," and not in_q:
                vals.append(cur.strip())
                cur = ""
            else:
                cur += ch
        vals.append(cur.strip())
        while len(vals) < len(header):
            vals.append("")
        return {header[i]: vals[i] for i in range(len(header))}

    rows = [parse_line(line) for line in lines[1:] if line.strip()]
    return header, rows


def parse_salary_csv(filepath):
    header, rows = read_csv_robust(filepath)
    cleaned = [h.strip() for h in header]
    missing = [c for c in SALARY_COLUMNS if c not in cleaned]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    records = []
    for row in rows:
        rec = {col: row.get(col, "").strip() for col in header}
        rec["_basic"] = parse_currency(rec["Basic Salary"])
        rec["_medical"] = parse_currency(rec["Medical Allowance"])
        rec["_house_rent"] = parse_currency(rec["House Rent"])
        rec["_fuel"] = parse_currency(rec["Fuel"])
        rec["_hp"] = parse_currency(rec["HP"])
        rec["_eb"] = parse_currency(rec["EB"])
        rec["_pf"] = parse_currency(rec["PF D/LE"])
        rec["_epfc"] = parse_currency(rec["EPFC"])
        rec["_adv"] = parse_currency(rec["ADV."])
        rec["_ded"] = parse_currency(rec["DED."])
        rec["_wht"] = parse_currency(rec["WHT"])
        rec["_eobi"] = parse_currency(rec["EOBI"])
        rec["_gt"] = parse_currency(rec["GT"])
        rec["_td"] = parse_currency(rec["TD"])
        rec["_payable"] = parse_currency(rec["Payable"])
        rec["_target"] = parse_currency(rec["Target Salary"])
        rec["_month_sort"] = parse_month_key(rec["Month"])
        records.append(rec)
    return records





def parse_salary_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty")
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    missing = [c for c in SALARY_COLUMNS if c not in header]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    records = []
    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        rec = {}
        for i, col in enumerate(header):
            val = row[i] if i < len(row) else ""
            rec[col] = str(val).strip() if val is not None else ""
        for col in SALARY_COLUMNS:
            if col not in rec:
                rec[col] = ""
        rec["_basic"] = parse_currency(rec["Basic Salary"])
        rec["_medical"] = parse_currency(rec["Medical Allowance"])
        rec["_house_rent"] = parse_currency(rec["House Rent"])
        rec["_fuel"] = parse_currency(rec["Fuel"])
        rec["_hp"] = parse_currency(rec["HP"])
        rec["_eb"] = parse_currency(rec["EB"])
        rec["_pf"] = parse_currency(rec["PF D/LE"])
        rec["_epfc"] = parse_currency(rec["EPFC"])
        rec["_adv"] = parse_currency(rec["ADV."])
        rec["_ded"] = parse_currency(rec["DED."])
        rec["_wht"] = parse_currency(rec["WHT"])
        rec["_eobi"] = parse_currency(rec["EOBI"])
        rec["_gt"] = parse_currency(rec["GT"])
        rec["_td"] = parse_currency(rec["TD"])
        rec["_payable"] = parse_currency(rec["Payable"])
        rec["_target"] = parse_currency(rec["Target Salary"])
        rec["_month_sort"] = parse_month_key(rec["Month"])
        records.append(rec)
    wb.close()
    return records


def _emp_from_record(rec):
    return {
        "_cnic": rec.get("CNIC", ""),
        "_phone": rec.get("Phone", ""),
        "_email": rec.get("Email", ""),
        "_bank": rec.get("Bank Name", ""),
        "_acct_title": rec.get("Account Title", ""),
        "_acct_number": rec.get("Account Number", ""),
    }


def build_slip_data(records, employee_details=None):
    months = OrderedDict()
    for rec in records:
        mk = rec["Month"].strip()
        if mk not in months:
            months[mk] = []
        r = dict(rec)
        r.update(_emp_from_record(rec))
        name_key = r["Name"].strip().lower()
        if employee_details and name_key in employee_details:
            ed = employee_details[name_key]
            for k in ("_cnic", "_phone", "_email", "_bank", "_acct_title", "_acct_number"):
                if ed.get(k):
                    r[k] = ed[k]
        months[mk].append(r)

    sorted_months = OrderedDict()
    for mk in sorted(months.keys(), key=lambda x: parse_month_key(x)):
        sorted_months[mk] = sorted(months[mk], key=lambda r: r["Name"].strip().lower())
    return sorted_months


def sanitize_filename(name):
    return re.sub(r'[<>:"/\\|?*]', '_', name.strip())


# ─── PDF generation: two professional payslips per A4 page ───────────────

LM, RM, TM, BM = 12*mm, 12*mm, 10*mm, 10*mm
W = A4[0] - LM - RM
H = A4[1] - TM - BM
G = 5*mm
SH = (H - G) / 2

PRI = "#1a3a5c"
ACC = "#2c6fbb"
GRN = "#1b7a3d"
RED = "#c62828"
BG = "#f8f9fa"
BG2 = "#edf2f7"
BD = "#d0d5dd"


class TwoUp(BaseDocTemplate):
    def __init__(self, buf, **kw):
        BaseDocTemplate.__init__(self, buf, **kw)
        ft = Frame(LM, BM + SH + G, W, SH, id="t")
        fb = Frame(LM, BM, W, SH, id="b")
        self.addPageTemplates([PageTemplate(id="m", frames=[ft, fb])])


def slip(rec, w):
    el = []
    def P(t, s): return Paragraph(t, s)
    def ps(n, s, c="#111827", b=False, a=TA_LEFT):
        return ParagraphStyle(n, fontSize=s, leading=s+2, textColor=colors.HexColor(c),
                             fontName="Helvetica-Bold" if b else "Helvetica", alignment=a)
    def lbl(t): return P(t, ps("lbl", 8.5, "#6b7280"))
    def val(t): return P(t, ps("val", 9.5, "#111827"))
    def amt(n, b=False):
        return P(f"Rs. {format_currency(n)}" if n else "-",
                 ps("amt", 9, "#111827", b, TA_RIGHT))

    # ── Header ──
    el.append(P("Solarkon (Private) Limited", ps("co", 14, PRI, True, TA_CENTER)))
    el.append(P(f"Payslip \u2014 {format_month_display(rec['Month'])}",
               ps("pe", 11, ACC, False, TA_CENTER)))
    el.append(Spacer(1, 4*mm))

    # ── Employee details ──
    lw = 70
    rows = [
        ["Employee Name", rec["Name"].strip(), "Designation", rec["Designation"].strip()],
        ["CNIC", rec.get("_cnic", ""), "Contact", rec.get("_phone", "")],
        ["Bank", rec.get("_bank", ""), "A/C No", rec.get("_acct_number", "")],
        ["Email", rec.get("_email", ""), "A/C Title", rec.get("_acct_title", "")],
    ]

    dr = []
    for r in rows:
        dr.append([lbl(r[0]), val(r[1]),
                   lbl(r[2]) if r[2] else P("", ps("e", 7)),
                   val(r[3]) if r[3] else P("", ps("e", 7))])

    dt = Table(dr, colWidths=[lw, w/2-lw, lw, w/2-lw])
    dt.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("LEFTPADDING", (0,0), (-1,-1), 7),
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor(BG)),
        ("BOX", (0,0), (-1,-1), 0.5, colors.HexColor(BD)),
        ("INNERGRID", (0,0), (-1,-1), 0.25, colors.HexColor(BD)),
    ]))
    el.append(dt)
    el.append(Spacer(1, 4*mm))

    # ── Salary: Earnings (left) | Deductions (right) ──
    earn = [("Basic Salary", rec["_basic"]), ("Medical Allowance", rec["_medical"]),
            ("House Rent", rec["_house_rent"]), ("Fuel", rec["_fuel"])]
    ded = [("Advance Paid", rec["_adv"]), ("Deductions", rec["_ded"]),
           ("Withholding Tax", rec["_wht"])]
    if rec["_pf"] != 0:
        ded.insert(0, ("PF D/LE", rec["_pf"]))

    mr = max(len(earn), len(ded), 3)
    hdr_s = ps("h", 9, PRI, True)

    comb = [[P("EARNINGS", hdr_s), "", P("DEDUCTIONS", hdr_s), ""]]
    for i in range(mr):
        el_, ev_ = earn[i] if i < len(earn) else ("", 0)
        dl_, dv_ = ded[i] if i < len(ded) else ("", 0)
        comb.append([
            P(el_ if el_ else " ", ps("el", 8.5)),
            amt(ev_),
            P(dl_ if dl_ else " ", ps("dl", 8.5)),
            amt(dv_),
        ])

    comb.append([
        P("Gross Salary", ps("gt", 9, "#111827", True)),
        amt(rec["_gt"], b=True),
        P("Total Deductions", ps("td", 9, "#111827", True)),
        amt(rec["_td"], b=True),
    ])

    ct = Table(comb, colWidths=[w/4, w/4, w/4, w/4])
    cmds = [
        ("VALIGN", (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("LEFTPADDING", (0,0),(-1,-1), 7),
        ("RIGHTPADDING", (0,0),(-1,-1), 7),
        ("BACKGROUND", (0,0),(1,0), colors.HexColor(BG2)),
        ("BACKGROUND", (2,0),(3,0), colors.HexColor(BG2)),
        ("BACKGROUND", (0,-1),(1,-1), colors.HexColor(BG2)),
        ("BACKGROUND", (2,-1),(3,-1), colors.HexColor(BG)),
        ("BOX", (0,0),(1,-1), 0.5, colors.HexColor(BD)),
        ("BOX", (2,0),(3,-1), 0.5, colors.HexColor(BD)),
        ("LINEBELOW", (0,0),(3,0), 0.75, colors.HexColor(PRI)),
        ("LINEABOVE", (0,-1),(1,-1), 0.5, colors.HexColor("#111827")),
        ("LINEABOVE", (2,-1),(3,-1), 0.5, colors.HexColor("#111827")),
        ("LINEAFTER", (1,0),(1,-1), 0.25, colors.HexColor(BD)),
    ]
    for i in range(1, len(comb)-1):
        cmds.append(("BACKGROUND", (0,i),(1,i), colors.HexColor(BG) if i%2==0 else colors.HexColor("#ffffff")))
        cmds.append(("BACKGROUND", (2,i),(3,i), colors.HexColor(BG) if i%2==0 else colors.HexColor("#ffffff")))
    ct.setStyle(TableStyle(cmds))
    el.append(ct)
    el.append(Spacer(1, 4*mm))

    # ── Summary ──
    sd = [[
        P("Net Salary Payable", ps("nsp", 10, "#111827", True)),
        P(f"Rs. {format_currency(rec['_payable'])}", ps("nsv", 10, "#111827", False, TA_RIGHT)),
        P("Previous Balance", ps("pb", 10, "#6b7280")),
        P("Rs. 0", ps("pbv", 10, "#6b7280", False, TA_RIGHT)),
        P("Net Payable", ps("np", 12, "#111827", True)),
        P(f"Rs. {format_currency(rec['_payable'])}", ps("npv", 12, "#111827", True, TA_RIGHT)),
    ]]
    st = Table(sd, colWidths=[w*0.22, w*0.14, w*0.16, w*0.12, w*0.20, w*0.16])
    st.setStyle(TableStyle([
        ("VALIGN", (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0),(-1,-1), 6),
        ("BOTTOMPADDING", (0,0),(-1,-1), 6),
        ("LEFTPADDING", (0,0),(-1,-1), 7),
        ("RIGHTPADDING", (0,0),(-1,-1), 7),
        ("BACKGROUND", (0,0),(-1,-1), colors.HexColor(BG)),
        ("BOX", (0,0),(-1,-1), 0.5, colors.HexColor(BD)),
        ("LINEAFTER", (1,0),(1,0), 0.25, colors.HexColor(BD)),
        ("LINEAFTER", (3,0),(3,0), 0.25, colors.HexColor(BD)),
        ("LINEAFTER", (4,0),(4,0), 0.25, colors.HexColor(BD)),
    ]))
    el.append(st)
    return el


def gen_pdf(slip_data, max_pgs=None):
    buf = io.BytesIO()
    doc = TwoUp(buf, pagesize=A4, leftMargin=LM, rightMargin=RM,
                topMargin=TM, bottomMargin=BM)
    all_r = [r for ms, rs in slip_data.items() for r in rs]
    total = len(all_r)
    if max_pgs:
        all_r = all_r[:max_pgs * 2]

    fls = []
    for i in range(0, len(all_r), 2):
        fls.extend(slip(all_r[i], W))
        if i + 1 < len(all_r):
            fls.append(FrameBreak())
            fls.extend(slip(all_r[i+1], W))
        if i + 2 < len(all_r):
            fls.append(PageBreak())

    doc.build(fls)
    buf.seek(0)
    return buf, total


def gen_separate(slip_data):
    bufs, names = [], []
    for ms, rs in slip_data.items():
        for r in rs:
            buf = io.BytesIO()
            doc = TwoUp(buf, pagesize=A4, leftMargin=LM, rightMargin=RM,
                        topMargin=TM, bottomMargin=BM)
            fls = slip(r, W)
            doc.build(fls)
            buf.seek(0)
            bufs.append(buf)
            names.append(f"{sanitize_filename(ms)}_{sanitize_filename(r['Name'].strip())}.pdf")
    return bufs, names


# ─── Template download ────────────────────────────────────────────────────


@app.route("/download_template")
def download_template():
    if not HAS_OPENPYXL:
        return jsonify({"error": "openpyxl is not installed. Unable to generate Excel template."}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salary Data"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a3a5c", end_color="1a3a5c", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_idx, col_name in enumerate(SALARY_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    sample_data = [
        ["Jan-2026", "John Doe", "Manager", 50000, 5000, 15000, 3000, 2000, 1000,
         "5000", 81000, 50000, 2500, 2000, 1500, 1000, 500, 1500, 72500, 80000],
        ["Jan-2026", "Jane Smith", "Assistant", 30000, 3000, 9000, 2000, 1000, 500,
         "3000", 48500, 30000, 1500, 1000, 800, 500, 500, 800, 43900, 48000],
    ]

    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    for col_idx, col_name in enumerate(SALARY_COLUMNS, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            max((len(str(row[col_idx-1])) for row in sample_data), default=0)
        )
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()

    return send_file(
        buf, as_attachment=True,
        download_name="Salary_Template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/download_employee_template")
def download_employee_template():
    if not HAS_OPENPYXL:
        return jsonify({"error": "openpyxl is not installed. Unable to generate Excel template."}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Details"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a3a5c", end_color="1a3a5c", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    cols = ["Name"] + EMPLOYEE_COLUMNS
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    sample_data = [
        ["John Doe", "42101-1234567-8", "0300-1234567", "john@example.com", "HBL", "John Doe", "0012-3456789-01"],
        ["Jane Smith", "42202-7654321-9", "0301-7654321", "jane@example.com", "UBL", "Jane Smith", "0098-7654321-02"],
    ]

    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    for col_idx in range(1, len(cols) + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()

    return send_file(
        buf, as_attachment=True,
        download_name="Employee_Template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ─── Flask routes ────────────────────────────────────────────────────────


@app.route("/")
def index():
    return render_template("index.html")


ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}


def _get_ext(filename):
    _, ext = os.path.splitext(filename)
    return ext.lower()


@app.route("/upload", methods=["POST"])
def upload():
    if "salary_csv" not in request.files:
        return jsonify({"error": "No salary file uploaded"}), 400
    salary_file = request.files["salary_csv"]
    if salary_file.filename == "":
        return jsonify({"error": "No file selected"}), 400
    ext = _get_ext(salary_file.filename)
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"error": "Please upload a CSV or Excel file (.csv, .xlsx, .xls)"}), 400
    if ext == ".xls":
        return jsonify({"error": "Legacy .xls format is not supported. Please save as .xlsx or .csv"}), 400

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    salary_path = os.path.join(
        app.config["UPLOAD_FOLDER"],
        f"salary_{ts}{ext}"
    )
    salary_file.save(salary_path)
    session["salary_path"] = salary_path

    try:
        if ext == ".csv":
            records = parse_salary_csv(salary_path)
        else:
            records = parse_salary_excel(salary_path)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Failed to parse file: {str(e)}"}), 400

    slip_data = build_slip_data(records)

    employees_set = sorted(set(r["Name"].strip().lower() for r in records))
    emp_details = {}
    for r in records:
        nk = r["Name"].strip().lower()
        if nk not in emp_details:
            emp_details[nk] = {
                "name": r["Name"].strip(),
                "cnic": r.get("CNIC", ""),
                "phone": r.get("Phone", ""),
                "email": r.get("Email", ""),
                "bank": r.get("Bank Name", ""),
                "acct_title": r.get("Account Title", ""),
                "acct_number": r.get("Account Number", ""),
            }

    preview = []
    for month_str, recs in slip_data.items():
        for rec in recs:
            preview.append({
                "month": rec["Month"].strip(),
                "name": rec["Name"].strip(),
                "designation": rec["Designation"].strip(),
                "basic": format_currency(rec["_basic"]),
                "medical": format_currency(rec["_medical"]),
                "house_rent": format_currency(rec["_house_rent"]),
                "fuel": format_currency(rec["_fuel"]),
                "gt": format_currency(rec["_gt"]),
                "adv": format_currency(rec["_adv"]),
                "td": format_currency(rec["_td"]),
                "payable": format_currency(rec["_payable"]),
                "cnic": rec.get("_cnic", ""),
                "phone": rec.get("_phone", ""),
                "email": rec.get("_email", ""),
                "bank": rec.get("_bank", ""),
                "acct_title": rec.get("_acct_title", ""),
                "acct_number": rec.get("_acct_number", ""),
            })

    months = list(slip_data.keys())
    total_employees = len(set(r["Name"].strip() for r in records))
    employees = sorted(set(r["Name"].strip() for r in records))
    designations = sorted(set(r["Designation"].strip() for r in records))
    years = sorted(set(m.split("-")[1] for m in months))

    return jsonify({
        "success": True,
        "preview": preview,
        "months": months,
        "years": years,
        "employees": employees,
        "designations": designations,
        "total_records": len(records),
        "total_employees": total_employees,
        "employee_details": [emp_details[nk] for nk in employees_set],
    })


@app.route("/save_employee_details", methods=["POST"])
def save_employee_details():
    data = request.get_json(silent=True)
    if not data or "details" not in data:
        return jsonify({"error": "No details provided"}), 400
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    path = os.path.join(app.config["UPLOAD_FOLDER"], f"emp_details_{ts}.json")
    import json as _json
    with open(path, "w", encoding="utf-8") as f:
        _json.dump(data["details"], f, indent=2)
    session["emp_details_path"] = path
    return jsonify({"success": True})


@app.route("/upload_employee_file", methods=["POST"])
def upload_employee_file():
    if "emp_file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["emp_file"]
    if f.filename == "":
        return jsonify({"error": "No file selected"}), 400
    ext = _get_ext(f.filename)
    if ext not in {".csv", ".xlsx"}:
        return jsonify({"error": "Please upload a CSV or Excel file (.csv, .xlsx)"}), 400
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    path = os.path.join(app.config["UPLOAD_FOLDER"], f"emp_upload_{ts}{ext}")
    f.save(path)
    try:
        if ext == ".csv":
            header, rows = read_csv_robust(path)
        else:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(h).strip() if h is not None else "" for h in rows[0]]
                rows = rows[1:]
            else:
                header, rows = [], []
            wb.close()
    except Exception as e:
        return jsonify({"error": f"Failed to parse file: {str(e)}"}), 400

    details = {}
    for row in rows:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        rec = {}
        for i, col in enumerate(header):
            val = row[i] if i < len(row) else ""
            rec[col.strip()] = str(val).strip() if val is not None else ""
        name = rec.get("Name", "").strip().lower()
        if name:
            details[name] = {
                "_cnic": rec.get("CNIC", ""),
                "_phone": rec.get("Phone", ""),
                "_email": rec.get("Email", ""),
                "_bank": rec.get("Bank Name", "") or rec.get("Bank Account", ""),
                "_acct_title": rec.get("Account Title", ""),
                "_acct_number": rec.get("Account Number", ""),
            }
    return jsonify({"success": True, "details": details})


@app.route("/preview_pdf", methods=["POST"])
def preview_pdf():
    salary_path = session.get("salary_path")
    if not salary_path or not os.path.exists(salary_path):
        return jsonify({"error": "No data. Please upload again."}), 400

    filters = request.get_json(silent=True) or {}
    selected_months = filters.get("months", [])
    selected_employees = filters.get("employees", [])
    selected_designations = filters.get("designations", [])

    import json as _json
    employee_details = None
    emp_path = session.get("emp_details_path")
    if emp_path and os.path.exists(emp_path):
        with open(emp_path, "r", encoding="utf-8") as f:
            employee_details = _json.load(f)

    try:
        ext = _get_ext(salary_path)
        if ext == ".csv":
            records = parse_salary_csv(salary_path)
        else:
            records = parse_salary_excel(salary_path)
    except Exception as e:
        return jsonify({"error": str(e)}), 400

    if selected_months:
        records = [r for r in records if r["Month"].strip() in selected_months]
    if selected_employees:
        records = [r for r in records if r["Name"].strip() in selected_employees]
    if selected_designations:
        records = [r for r in records if r["Designation"].strip() in selected_designations]

    slip_data = build_slip_data(records, employee_details)
    pdf_buffer, total = gen_pdf(slip_data)

    return send_file(
        pdf_buffer, mimetype="application/pdf",
        as_attachment=False, download_name="preview.pdf"
    )


@app.route("/generate", methods=["POST"])
def generate():
    data = request.get_json()
    output_type = data.get("type", "combined")
    selected_months = data.get("months", [])
    selected_employees = data.get("employees", [])
    selected_designations = data.get("designations", [])

    salary_path = session.get("salary_path")
    if not salary_path or not os.path.exists(salary_path):
        return jsonify({"error": "No uploaded data found. Please upload again."}), 400

    import json as _json
    employee_details = None
    emp_path = session.get("emp_details_path")
    if emp_path and os.path.exists(emp_path):
        with open(emp_path, "r", encoding="utf-8") as f:
            employee_details = _json.load(f)

    try:
        ext = _get_ext(salary_path)
        if ext == ".csv":
            records = parse_salary_csv(salary_path)
        else:
            records = parse_salary_excel(salary_path)
    except Exception as e:
        return jsonify({"error": f"Failed to parse file: {str(e)}"}), 400

    if selected_months:
        records = [r for r in records if r["Month"].strip() in selected_months]
    if selected_employees:
        records = [r for r in records if r["Name"].strip() in selected_employees]
    if selected_designations:
        records = [r for r in records if r["Designation"].strip() in selected_designations]

    slip_data = build_slip_data(records, employee_details)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if output_type == "combined":
        pdf_buffer, total = gen_pdf(slip_data)
        filename = f"Payslips_Combined_{timestamp}.pdf"
        return send_file(
            pdf_buffer, as_attachment=True,
            download_name=filename, mimetype="application/pdf"
        )

    elif output_type == "separate":
        buffers, names = gen_separate(slip_data)
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for buf, name in zip(buffers, names):
                zf.writestr(name, buf.read())
        zip_buffer.seek(0)
        filename = f"Payslips_Separate_{timestamp}.zip"
        return send_file(
            zip_buffer, as_attachment=True,
            download_name=filename, mimetype="application/zip"
        )

    return jsonify({"error": "Invalid output type"}), 400


@app.route("/cleanup", methods=["POST"])
def cleanup():
    for folder in [app.config["UPLOAD_FOLDER"], app.config["OUTPUT_FOLDER"]]:
        if os.path.exists(folder):
            for f in os.listdir(folder):
                fp = os.path.join(folder, f)
                if os.path.isfile(fp):
                    try:
                        os.remove(fp)
                    except Exception:
                        pass
    return jsonify({"success": True})


if __name__ == "__main__":
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("DEBUG", "0") == "1"

    import threading
    threading.Timer(1.5, lambda: __import__("webbrowser").open(f"http://{host}:{port}")).start()

    app.run(host=host, port=port, debug=debug)
