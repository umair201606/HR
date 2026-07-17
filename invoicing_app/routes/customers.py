import json
from io import BytesIO
from flask import Blueprint, render_template, redirect, url_for, flash, request, session, send_file
from flask_login import login_required
from inventory_app.extensions import db
from inventory_app.models.customer import InvCustomer
from shared.ledger_utils import create_entity_account
from shared.permissions import deny_page

inv_cust_bp = Blueprint("inv_customers", __name__, url_prefix="/inventory/customers")


@inv_cust_bp.route("/")
@login_required
def list_customers():
    q = request.args.get("q", "")
    query = InvCustomer.query
    if q:
        query = query.filter(
            InvCustomer.name.ilike(f"%{q}%") | InvCustomer.city.ilike(f"%{q}%")
        )
    customers = query.order_by(InvCustomer.name).all()
    return render_template("customers/list_inv.html", customers=customers)


@inv_cust_bp.route("/create", methods=["GET", "POST"])
@login_required
def create_customer():
    if deny_page("customers", "create"):
        return redirect(url_for("inv_customers.list_customers"))
    if request.method == "POST":
        c = InvCustomer(
            name=request.form["name"],
            contact_person=request.form.get("contact_person", ""),
            email=request.form.get("email", ""),
            phone=request.form.get("phone", ""),
            mobile=request.form.get("mobile", ""),
            address=request.form.get("address", ""),
            city=request.form.get("city", ""),
            tax_id=request.form.get("tax_id", ""),
            payment_terms=request.form.get("payment_terms", ""),
            credit_limit=request.form.get("credit_limit", 0, type=float),
            website=request.form.get("website", ""),
            notes=request.form.get("notes", ""),
        )
        db.session.add(c)
        db.session.flush()
        create_entity_account("customer", c.id, c.name)
        db.session.commit()
        flash(f"Customer created — ledger account '{c.name}' added under Trade Debtors", "success")
        return redirect(url_for("inv_customers.list_customers"))
    return render_template("customers/form_inv.html", customer=None)


@inv_cust_bp.route("/edit/<int:id>", methods=["GET", "POST"])
@login_required
def edit_customer(id):
    if deny_page("customers", "edit"):
        return redirect(url_for("inv_customers.list_customers"))
    c = InvCustomer.query.get_or_404(id)
    if request.method == "POST":
        c.name = request.form["name"]
        c.contact_person = request.form.get("contact_person", "")
        c.email = request.form.get("email", "")
        c.phone = request.form.get("phone", "")
        c.mobile = request.form.get("mobile", "")
        c.address = request.form.get("address", "")
        c.city = request.form.get("city", "")
        c.tax_id = request.form.get("tax_id", "")
        c.payment_terms = request.form.get("payment_terms", "")
        c.credit_limit = request.form.get("credit_limit", 0, type=float)
        c.website = request.form.get("website", "")
        c.notes = request.form.get("notes", "")
        c.is_active = request.form.get("is_active") == "on"
        create_entity_account("customer", c.id, c.name)
        db.session.commit()
        flash("Customer updated", "success")
        return redirect(url_for("inv_customers.list_customers"))
    return render_template("customers/form_inv.html", customer=c)


@inv_cust_bp.route("/delete/<int:id>")
@login_required
def delete_customer(id):
    c = InvCustomer.query.get_or_404(id)
    if c.sales_orders.count() > 0 or c.invoices.count() > 0:
        flash("Cannot delete customer with sales history", "error")
    else:
        db.session.delete(c)
        db.session.commit()
        flash("Customer deleted", "success")
    return redirect(url_for("inv_customers.list_customers"))


@inv_cust_bp.route("/check")
@login_required
def check_duplicate():
    value = request.args.get("value", "").strip()
    exclude = request.args.get("exclude", type=int)
    if not value:
        return {"exists": False}
    q = InvCustomer.query.filter(InvCustomer.name == value)
    if exclude:
        q = q.filter(InvCustomer.id != exclude)
    exists = q.first() is not None
    return {"exists": exists}


@inv_cust_bp.route("/template")
@login_required
def download_template():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    ws.append(["Name", "Contact Person", "Email", "Phone", "Mobile", "City", "Tax ID", "Payment Terms", "Credit Limit", "Website", "Address", "Notes"])
    ws.append(["Sample Customer A", "John Doe", "john@example.com", "021-1234567", "0300-1234567", "Karachi", "NTN-1234567", "Net 30", "50000", "https://example.com", "123 Main St", "Regular customer"])
    ws.append(["Sample Customer B", "Jane Smith", "jane@example.com", "042-7654321", "0301-7654321", "Lahore", "", "Due on Receipt", "25000", "", "456 Oak Ave", ""])
    ws.append(["Sample Customer C", "", "", "", "", "Islamabad", "", "Net 15", "10000", "", "", ""])
    for col, w in enumerate([28, 18, 24, 14, 14, 14, 16, 16, 12, 22, 24, 24], start=1):
        ws.column_dimensions[chr(64 + col)].width = w
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name="customer_import_template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@inv_cust_bp.route("/upload-excel", methods=["POST"])
@login_required
def upload_excel():
    file = request.files.get("file")
    if not file:
        flash("No file selected.", "error")
        return redirect(url_for("inv_customers.list_customers"))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
        ws = wb.active
        if not ws:
            flash("Excel file is empty.", "error")
            return redirect(url_for("inv_customers.list_customers"))

        rows = []
        errors = []
        seen_names = set()

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            name = (str(row[0] or "").strip()) if row[0] is not None else ""
            contact = (str(row[1] or "").strip()) if len(row) > 1 and row[1] is not None else ""
            email = (str(row[2] or "").strip()) if len(row) > 2 and row[2] is not None else ""
            phone = (str(row[3] or "").strip()) if len(row) > 3 and row[3] is not None else ""
            mobile = (str(row[4] or "").strip()) if len(row) > 4 and row[4] is not None else ""
            city = (str(row[5] or "").strip()) if len(row) > 5 and row[5] is not None else ""
            tax_id = (str(row[6] or "").strip()) if len(row) > 6 and row[6] is not None else ""
            payment_terms = (str(row[7] or "").strip()) if len(row) > 7 and row[7] is not None else ""
            credit_limit = 0.0
            try:
                if len(row) > 8 and row[8] is not None:
                    credit_limit = float(str(row[8]).replace(",", ""))
            except (ValueError, TypeError):
                errors.append(f"Row {i}: Invalid Credit Limit.")
            website = (str(row[9] or "").strip()) if len(row) > 9 and row[9] is not None else ""
            address = (str(row[10] or "").strip()) if len(row) > 10 and row[10] is not None else ""
            notes = (str(row[11] or "").strip()) if len(row) > 11 and row[11] is not None else ""

            if not name:
                continue

            if name in seen_names:
                errors.append(f"Row {i}: Duplicate name '{name}' in spreadsheet.")
                continue

            existing = InvCustomer.query.filter_by(name=name).first()
            if existing:
                errors.append(f"Row {i}: '{name}' already exists in database.")
                continue

            seen_names.add(name)
            rows.append({
                "name": name, "contact_person": contact, "email": email,
                "phone": phone, "mobile": mobile, "city": city,
                "tax_id": tax_id, "payment_terms": payment_terms,
                "credit_limit": credit_limit, "website": website,
                "address": address, "notes": notes,
            })

        if not rows:
            flash("No valid rows found to import." + (" " + "; ".join(errors) if errors else ""), "error")
            return redirect(url_for("inv_customers.list_customers"))

        session["batch_customers"] = rows
        flash(f"Parsed {len(rows)} customer(s) from Excel." + (" Warnings: " + "; ".join(errors) if errors else ""), "success" if not errors else "warning")
        return redirect(url_for("inv_customers.batch_editor"))

    except Exception as e:
        flash(f"Failed to parse Excel file: {e}", "error")
        return redirect(url_for("inv_customers.list_customers"))


@inv_cust_bp.route("/batch", methods=["GET", "POST"])
@login_required
def batch_editor():
    if request.method == "POST":
        raw = request.form.get("data", "[]")
        try:
            rows = json.loads(raw)
        except (json.JSONDecodeError, TypeError):
            flash("Invalid data submitted.", "error")
            return redirect(url_for("inv_customers.list_customers"))

        created = 0
        errors = []
        for r in rows:
            name = r.get("name", "").strip()
            row_marked = r.get("_delete", False)
            if row_marked:
                continue
            if not name:
                errors.append("Row with empty name skipped.")
                continue

            if InvCustomer.query.filter_by(name=name).first():
                errors.append(f"'{name}' — already exists.")
                continue

            try:
                credit_limit = float(r.get("credit_limit", 0))
            except (ValueError, TypeError):
                credit_limit = 0

            c = InvCustomer(
                name=name,
                contact_person=r.get("contact_person", "").strip(),
                email=r.get("email", "").strip(),
                phone=r.get("phone", "").strip(),
                mobile=r.get("mobile", "").strip(),
                address=r.get("address", "").strip(),
                city=r.get("city", "").strip(),
                tax_id=r.get("tax_id", "").strip(),
                payment_terms=r.get("payment_terms", "").strip(),
                credit_limit=credit_limit,
                website=r.get("website", "").strip(),
                notes=r.get("notes", "").strip(),
            )
            db.session.add(c)
            db.session.flush()
            create_entity_account("customer", c.id, c.name)
            created += 1

        db.session.commit()
        session.pop("batch_customers", None)
        msg = f"{created} customer(s) created."
        if errors:
            msg += " Errors: " + "; ".join(errors)
        flash(msg, "success" if not errors else "warning")
        return redirect(url_for("inv_customers.list_customers"))

    rows = session.pop("batch_customers", [])
    return render_template("customers/batch_editor.html", rows=rows)
