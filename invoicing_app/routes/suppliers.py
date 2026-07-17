import json
from io import BytesIO
from flask import Blueprint, render_template, redirect, url_for, flash, request, session, send_file
from flask_login import login_required
from inventory_app.extensions import db
from inventory_app.models.supplier import InvSupplier
from shared.ledger_utils import create_entity_account
from shared.permissions import deny_page

inv_sup_bp = Blueprint("inv_suppliers", __name__, url_prefix="/inventory/suppliers")


@inv_sup_bp.route("/")
@login_required
def list_suppliers():
    q = request.args.get("q", "")
    query = InvSupplier.query
    if q:
        query = query.filter(
            InvSupplier.name.ilike(f"%{q}%") | InvSupplier.city.ilike(f"%{q}%")
        )
    suppliers = query.order_by(InvSupplier.name).all()
    return render_template("suppliers/list_inv.html", suppliers=suppliers)


@inv_sup_bp.route("/create", methods=["GET", "POST"])
@login_required
def create_supplier():
    if deny_page("suppliers", "create"):
        return redirect(url_for("inv_suppliers.list_suppliers"))
    if request.method == "POST":
        s = InvSupplier(
            name=request.form["name"],
            contact_person=request.form.get("contact_person", ""),
            email=request.form.get("email", ""),
            phone=request.form.get("phone", ""),
            mobile=request.form.get("mobile", ""),
            address=request.form.get("address", ""),
            city=request.form.get("city", ""),
            tax_id=request.form.get("tax_id", ""),
            payment_terms=request.form.get("payment_terms", ""),
            website=request.form.get("website", ""),
            notes=request.form.get("notes", ""),
        )
        db.session.add(s)
        db.session.flush()
        create_entity_account("supplier", s.id, s.name)
        db.session.commit()
        flash(f"Supplier created — ledger account '{s.name}' added under Trade Creditors", "success")
        return redirect(url_for("inv_suppliers.list_suppliers"))
    return render_template("suppliers/form_inv.html", supplier=None)


@inv_sup_bp.route("/edit/<int:id>", methods=["GET", "POST"])
@login_required
def edit_supplier(id):
    if deny_page("suppliers", "edit"):
        return redirect(url_for("inv_suppliers.list_suppliers"))
    s = InvSupplier.query.get_or_404(id)
    if request.method == "POST":
        s.name = request.form["name"]
        s.contact_person = request.form.get("contact_person", "")
        s.email = request.form.get("email", "")
        s.phone = request.form.get("phone", "")
        s.mobile = request.form.get("mobile", "")
        s.address = request.form.get("address", "")
        s.city = request.form.get("city", "")
        s.tax_id = request.form.get("tax_id", "")
        s.payment_terms = request.form.get("payment_terms", "")
        s.website = request.form.get("website", "")
        s.notes = request.form.get("notes", "")
        s.is_active = request.form.get("is_active") == "on"
        create_entity_account("supplier", s.id, s.name)
        db.session.commit()
        flash("Supplier updated", "success")
        return redirect(url_for("inv_suppliers.list_suppliers"))
    return render_template("suppliers/form_inv.html", supplier=s)


@inv_sup_bp.route("/delete/<int:id>")
@login_required
def delete_supplier(id):
    if deny_page("suppliers", "delete"):
        return redirect(url_for("inv_suppliers.list_suppliers"))
    s = InvSupplier.query.get_or_404(id)
    if s.purchase_orders.count() > 0:
        flash("Cannot delete supplier with purchase orders", "error")
    else:
        db.session.delete(s)
        db.session.commit()
        flash("Supplier deleted", "success")
    return redirect(url_for("inv_suppliers.list_suppliers"))


@inv_sup_bp.route("/check")
@login_required
def check_duplicate():
    value = request.args.get("value", "").strip()
    exclude = request.args.get("exclude", type=int)
    if not value:
        return {"exists": False}
    q = InvSupplier.query.filter(InvSupplier.name == value)
    if exclude:
        q = q.filter(InvSupplier.id != exclude)
    exists = q.first() is not None
    return {"exists": exists}


@inv_sup_bp.route("/template")
@login_required
def download_template():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suppliers"
    ws.append(["Name", "Contact Person", "Email", "Phone", "Mobile", "City", "Tax ID", "Payment Terms", "Website", "Address", "Notes"])
    ws.append(["Sample Supplier A", "Ali Khan", "ali@supplier.com", "021-1112222", "0300-1112222", "Karachi", "NTN-7654321", "Net 30", "https://supplier-a.com", "789 Industrial Rd", ""])
    ws.append(["Sample Supplier B", "Sara Ahmed", "sara@supplier.com", "042-3334444", "0301-3334444", "Lahore", "", "Net 45", "", "321 Trade Ave", "Preferred vendor"])
    ws.append(["Sample Supplier C", "", "", "", "", "Islamabad", "", "Due on Receipt", "", "", ""])
    for col, w in enumerate([28, 18, 24, 14, 14, 14, 16, 16, 22, 24, 24], start=1):
        ws.column_dimensions[chr(64 + col)].width = w
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name="supplier_import_template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@inv_sup_bp.route("/upload-excel", methods=["POST"])
@login_required
def upload_excel():
    file = request.files.get("file")
    if not file:
        flash("No file selected.", "error")
        return redirect(url_for("inv_suppliers.list_suppliers"))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
        ws = wb.active
        if not ws:
            flash("Excel file is empty.", "error")
            return redirect(url_for("inv_suppliers.list_suppliers"))

        rows = []
        errors = []
        seen_names = set()

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            name = (str(row[0] or "").strip()) if row[0] is not None else ""
            contact = (str(row[1] or "").strip()) if len(row) > 1 and row[1] is not None else ""
            email = (str(row[2] or "").strip()) if len(row) > 2 and row[2] is not None else ""
            phone = (str(row[3] or "").strip()) if len(row) > 3 and row[3] is not None else ""
            mobile = (str(row[4] or "").strip()) if len(row) > 4 and row[4] is not None else ""
            city = (str(row[5] or "").strip()) if len(row) > 5 and row[5] is not None else ""
            tax_id = (str(row[6] or "").strip()) if len(row) > 6 and row[6] is not None else ""
            payment_terms = (str(row[7] or "").strip()) if len(row) > 7 and row[7] is not None else ""
            website = (str(row[8] or "").strip()) if len(row) > 8 and row[8] is not None else ""
            address = (str(row[9] or "").strip()) if len(row) > 9 and row[9] is not None else ""
            notes = (str(row[10] or "").strip()) if len(row) > 10 and row[10] is not None else ""

            if not name:
                continue

            if name in seen_names:
                errors.append(f"Row {i}: Duplicate name '{name}' in spreadsheet.")
                continue

            existing = InvSupplier.query.filter_by(name=name).first()
            if existing:
                errors.append(f"Row {i}: '{name}' already exists in database.")
                continue

            seen_names.add(name)
            rows.append({
                "name": name, "contact_person": contact, "email": email,
                "phone": phone, "mobile": mobile, "city": city,
                "tax_id": tax_id, "payment_terms": payment_terms,
                "website": website, "address": address, "notes": notes,
            })

        if not rows:
            flash("No valid rows found to import." + (" " + "; ".join(errors) if errors else ""), "error")
            return redirect(url_for("inv_suppliers.list_suppliers"))

        session["batch_suppliers"] = rows
        flash(f"Parsed {len(rows)} supplier(s) from Excel." + (" Warnings: " + "; ".join(errors) if errors else ""), "success" if not errors else "warning")
        return redirect(url_for("inv_suppliers.batch_editor"))

    except Exception as e:
        flash(f"Failed to parse Excel file: {e}", "error")
        return redirect(url_for("inv_suppliers.list_suppliers"))


@inv_sup_bp.route("/batch", methods=["GET", "POST"])
@login_required
def batch_editor():
    if request.method == "POST":
        raw = request.form.get("data", "[]")
        try:
            rows = json.loads(raw)
        except (json.JSONDecodeError, TypeError):
            flash("Invalid data submitted.", "error")
            return redirect(url_for("inv_suppliers.list_suppliers"))

        created = 0
        errors = []
        for r in rows:
            name = r.get("name", "").strip()
            row_marked = r.get("_delete", False)
            if row_marked:
                continue
            if not name:
                errors.append("Row with empty name skipped.")
                continue

            if InvSupplier.query.filter_by(name=name).first():
                errors.append(f"'{name}' — already exists.")
                continue

            s = InvSupplier(
                name=name,
                contact_person=r.get("contact_person", "").strip(),
                email=r.get("email", "").strip(),
                phone=r.get("phone", "").strip(),
                mobile=r.get("mobile", "").strip(),
                address=r.get("address", "").strip(),
                city=r.get("city", "").strip(),
                tax_id=r.get("tax_id", "").strip(),
                payment_terms=r.get("payment_terms", "").strip(),
                website=r.get("website", "").strip(),
                notes=r.get("notes", "").strip(),
            )
            db.session.add(s)
            db.session.flush()
            create_entity_account("supplier", s.id, s.name)
            created += 1

        db.session.commit()
        session.pop("batch_suppliers", None)
        msg = f"{created} supplier(s) created."
        if errors:
            msg += " Errors: " + "; ".join(errors)
        flash(msg, "success" if not errors else "warning")
        return redirect(url_for("inv_suppliers.list_suppliers"))

    rows = session.pop("batch_suppliers", [])
    return render_template("suppliers/batch_editor.html", rows=rows)
